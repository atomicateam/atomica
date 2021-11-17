# -*- coding: utf-8 -*-
"""
Miscellaneous utility functions for Excel files

This module implements utility functions for Excel functionality that is common
to different kinds of spreadsheets used in Atomica (e.g. Databooks and Program Books).
For example, Excel formatting, and time-varying data entry tables, are implemented here.

"""

from xlsxwriter.utility import xl_rowcol_to_cell as xlrc
import sciris as sc
import io
import numpy as np
from .system import FrameworkSettings as FS
import pandas as pd
from .utils import format_duration, datetime_to_year
import xlsxwriter

# Suppress known warning in Openpyxl
# Warnings are:
# - C:\ProgramData\Miniconda3\envs\atomica37\lib\site-packages\openpyxl\worksheet\_reader.py:300: UserWarning: Conditional Formatting extension is not supported and will be removed
#   warn(msg)
# - C:\ProgramData\Miniconda3\envs\atomica37\lib\site-packages\openpyxl\worksheet\_reader.py:300: UserWarning: Data Validation extension is not supported and will be removed
#   warn(msg)
# This means that conditional formatting and data valuation rules aren't being loaded, but since `data_only=True` these don't matter and can be safely ignored
import warnings

warnings.filterwarnings(action="ignore", category=UserWarning, module="openpyxl.worksheet", lineno=300)


def standard_formats(workbook):
    # Add standard formatting to a workbook and return the set of format objects
    # for use when writing within the workbook
    """the formats used in the spreadsheet"""
    #    darkgray = '#413839'
    #    optima_blue = '#18C1FF'
    atomica_blue = "#98E0FA"
    optional_orange = "#FFA500"
    BG_COLOR = atomica_blue
    OPT_COLOR = optional_orange
    BORDER_COLOR = "white"

    formats = {}

    # Locked formats
    formats["bold"] = workbook.add_format({"bold": 1})
    formats["center"] = workbook.add_format({"align": "center"})
    formats["center_bold"] = workbook.add_format({"bold": 1, "align": "center"})
    formats["rc_title"] = {}
    formats["rc_title"]["right"] = {}
    formats["rc_title"]["right"]["T"] = workbook.add_format({"bold": 1, "align": "right", "text_wrap": True})
    formats["rc_title"]["right"]["F"] = workbook.add_format({"bold": 1, "align": "right", "text_wrap": False})
    formats["rc_title"]["left"] = {}
    formats["rc_title"]["left"]["T"] = workbook.add_format({"bold": 1, "align": "left", "text_wrap": True})
    formats["rc_title"]["left"]["F"] = workbook.add_format({"bold": 1, "align": "left", "text_wrap": False})

    # Unlocked formats
    formats["unlocked"] = workbook.add_format({"locked": 0, "bg_color": BG_COLOR, "border": 1, "border_color": BORDER_COLOR})
    formats["center_unlocked"] = workbook.add_format({"align": "center", "locked": 0, "bg_color": BG_COLOR, "border": 1, "border_color": BORDER_COLOR})
    formats["general"] = workbook.add_format({"locked": 0, "num_format": 0x00, "bg_color": BG_COLOR, "border": 1, "border_color": BORDER_COLOR})

    # Conditional formats
    formats["unlocked_boolean_true"] = workbook.add_format({"bg_color": OPT_COLOR})
    formats["unlocked_boolean_false"] = workbook.add_format({"bg_color": BG_COLOR})
    formats["not_required"] = workbook.add_format({"bg_color": "#EEEEEE", "border": 1, "border_color": "#CCCCCC"})
    formats["white_bg"] = workbook.add_format({"bg_color": "#FFFFFF", "border": 1, "border_color": "#CCCCCC"})
    formats["ignored"] = workbook.add_format({"pattern": 14})  # Hatched with diagonal lines - this represents a cell whose value will not be used in the model run (e.g., an assumption that also has time-specific points)
    formats["warning"] = workbook.add_format({"bg_color": "#FF0000"})
    formats["ignored_warning"] = workbook.add_format({"pattern": 14, "bg_color": "#FF0000"})  # hatched, with red background
    formats["ignored_not_required"] = workbook.add_format({"pattern": 14, "bg_color": "#EEEEEE", "border": 1, "border_color": "#CCCCCC"})  # hatched, with grey background

    return formats


def apply_widths(worksheet, width_dict):
    for idx, width in width_dict.items():
        worksheet.set_column(idx, idx, width * 1.1 + 1)


def update_widths(width_dict: dict, column_index: int, contents: str) -> None:
    """
    Keep track of required width for a column

    ``width_dict`` is a dict that is keyed by column index e.g. 0,1,2
    and the value is the length of the longest contents seen for that column

    :param width_dict: Storage dictionary
    :param column_index: Index of the column value has been inserted in
    :param contents: Content, length of which is used to set width

    """

    if width_dict is None or contents is None or not sc.isstring(contents):
        return

    if len(contents) == 0:
        return

    if column_index not in width_dict:
        width_dict[column_index] = len(contents)
    else:
        width_dict[column_index] = max(width_dict[column_index], len(contents))


def transfer_comments(target: sc.Spreadsheet, comment_source: sc.Spreadsheet) -> None:
    """
    Copy comments between spreadsheets

    This function copies comments from one spreadsheet to another. Under the hood,
    a new spreadsheet is created with values from the ``target`` Spreadsheet
    and cell-wise formatting from the ``comment_source`` Spreadsheet. If a cell exists in
    this spreadsheet and not in the source, it will be retained as-is. If more cells exist in
    the ``comment_source`` than in this spreadsheet, those cells will be dropped. If a sheet exists in
    the ``comment_source`` and not in the current workbook, it will be added

    :param target: The target spreadsheet to write comments into
    :param comment_source: The source spreadsheet containing comments

    """

    import openpyxl
    from openpyxl.comments import Comment

    assert isinstance(target, sc.Spreadsheet)
    assert isinstance(comment_source, sc.Spreadsheet)

    this_workbook = openpyxl.load_workbook(target.tofile(), data_only=False)  # This is the value source workbook
    old_workbook = openpyxl.load_workbook(comment_source.tofile(), data_only=False)  # A openpyxl workbook for the old content

    for sheet in this_workbook.worksheets:

        # If this sheet isn't in the old workbook, do nothing
        if sheet.title not in old_workbook.sheetnames:
            continue

        # Transfer comments
        for row in old_workbook[sheet.title].rows:
            for cell in row:
                if cell.comment:
                    sheet[cell.coordinate].comment = Comment(cell.comment.text, "")

    # Save the modified spreadsheet to a new buffer
    f = io.BytesIO()
    this_workbook.save(f)
    f.flush()
    f.seek(0)
    target.load(f)


def copy_sheet(source: str, sheet_name: str, workbook: xlsxwriter.Workbook) -> None:
    """
    Copy a sheet into a Workbook

    This function allows values to be copied from a file into a Workbook. The
    main use case is to support custom sheets in databooks that are not otherwise parsed
    but which might need to be retained. In particular, the ``ProjectData`` class does
    not parse ignored sheets at all, because no guarantees can be made about the quantity and
    type of the content, and whether there are formulas etc. that would be lost. In some cases
    though, it may be necessary to edit the databook and preserve specific sheets. In general,
    this can be done by using the ``to_workbook()`` method and then manually performing whatever
    operations are necessary to preserve the content on the extra sheets. However, when the
    extra sheet content is extremely simple e.g. just a table of values, then this helper
    function can be used to facilitate copying the content.

    Warning - note that Excel functions, formatting, and comments will NOT be preserved.

    :param source: File name of the spreadsheet to read the source sheet from or an ``sc.Spreadsheet`` instance
    :param sheet_name: Name of the sheet to write to
    :param workbook: A Workbook instance to add the sheet to
    :return: None - the sheet will be added to the Workbook in-place
    """

    import openpyxl

    if sc.isstring(source):
        source = sc.Spreadsheet(source)

    src_workbook = openpyxl.load_workbook(source.tofile(), read_only=True, data_only=True)  # Load in read-only mode for performance, since we don't parse comments etc.
    src_worksheet = src_workbook[sheet_name]
    dst_worksheet = workbook.add_worksheet(sheet_name)

    for i, row in enumerate(src_worksheet.rows):
        for j, cell in enumerate(row):
            dst_worksheet.write(i, j, cell.value)

    src_workbook.close()


def read_tables(worksheet) -> tuple:
    """
    Read tables from sheet

    :param worksheet: An openpyxl worksheet
    :return: A tuple containing - A list of tables (which is a list of rows, terminated by an empty row in the original spreadsheet),
             and a list of start row indices for each table read in
    """
    # This function takes in a openpyxl worksheet, and returns tables
    # A table consists of a block of rows with any #ignore rows skipped
    # This function will start at the top of the worksheet, read rows into a buffer
    # until it gets to the first entirely empty row
    # And then returns the contents of that buffer as a table. So a table is a list of openpyxl rows
    # This function continues until it has exhausted all of the rows in the sheet

    buffer = []
    tables = []
    start_rows = []
    start = None

    for i, row in enumerate(worksheet.rows):

        # Skip any rows starting with '#ignore'
        if len(row) > 0 and row[0].data_type == "s" and row[0].value.startswith("#ignore"):
            continue  # Move on to the next row if row skipping is marked True

        # Find out whether we need to add the row to the buffer
        for cell in row:
            if cell.value:  # If the row has a non-empty cell, add the row to the buffer
                if not buffer:
                    start = i + 1  # Excel rows are indexed starting at 1
                buffer.append(row)
                break
        else:  # If the row was empty, then yield the buffer and flag that it should be cleared at the next iteration
            if buffer:
                tables.append(buffer)  # Only append the buffer if it is not empty
                start_rows.append(start)
            buffer = []

    # After the last row, if the buffer has some un-flushed contents, then yield it
    if buffer:
        tables.append(buffer)
        start_rows.append(start)

    return tables, start_rows


def read_dataframes(worksheet, merge=False) -> list:
    """
    Read dataframes from sheet

    This function operates similarly to ``read_tables`` except it returns Dataframes instead of
    cells. This enables the dataframes to be constructed more quickly, at the expense of being
    able to track the cell references and row numbers. These are shown for databooks (via ``read_tables``)
    but not for frameworks (which go via ``read_dataframes``)


    :param worksheet: An openpyxl worksheet
    :param merge: If False (default) then blank rows will be used to split the dataframes. If True, only one
                  DataFrame will be returned
    :return: A list of DataFrames

    """
    # This function takes in a openpyxl worksheet, and returns tables
    # A table consists of a block of rows with any #ignore rows skipped
    # This function will start at the top of the worksheet, read rows into a buffer
    # until it gets to the first entirely empty row
    # And then returns the contents of that buffer as a table. So a table is a list of openpyxl rows
    # This function continues until it has exhausted all of the rows in the sheet

    content = np.empty((worksheet.max_row, worksheet.max_column), dtype="object")
    ignore = np.zeros((worksheet.max_row), dtype=bool)
    empty = np.zeros((worksheet.max_row), dtype=bool)  # True for index where a new table begins

    for i, row in enumerate(worksheet.rows):
        if len(row) > 0 and (row[0].data_type == "s" and row[0].value.startswith("#ignore")):
            ignore[i] = True
            continue

        any_values = False
        for j, cell in enumerate(row):
            v = cell.value
            try:
                v = v.strip()
                has_value = bool(v)  # If it's a string type, call strip() before checking truthiness
            except AttributeError:
                has_value = v is not None  # If it's not a string type, then only consider it empty if it's type is None (otherwise, a numerical value of 0 would be treated as empty)
            if has_value:
                any_values = True
            content[i, j] = v
        if not any_values:
            empty[i] = True

    tables = []
    if merge:
        ignore[empty] = True
        if all(ignore):
            return []
        tables.append(content[~ignore, :])
    else:
        # A change from False to True means that we need to start a new table
        # A True followed by a True doesn't start a new table but instead gets ignored
        content = content[~ignore, :]
        empty = empty[~ignore]

        # If there is no content at all, return immediately
        if all(empty):
            return []

        idx = []
        for i in range(len(empty) - 1):
            if not empty[i] and not idx:
                # Write the first line. This could be followed by an empty row, so need to a separate block for this
                idx.append(i)

            if not empty[i] and empty[i + 1]:
                # row i is the last row in the table (so need to include it in the range, hence +1)
                idx.append(i + 1)
            elif empty[i] and not empty[i + 1]:
                # Row i+1 marks the start of a table
                idx.append(i + 1)

        if not empty[-1]:
            # If the last row has content, then make sure that the last table goes all the way up
            idx.append(empty.size)

        assert not len(idx) % 2, "Error in table parsing routine, did not correctly identify table breaks"

        tables = []
        for i in range(0, len(idx) - 1, 2):
            tables.append(content[idx[i] : idx[i + 1]].copy())

    dfs = []
    for table in tables:
        df = pd.DataFrame(table)
        df.dropna(axis=1, how="all", inplace=True)
        df.columns = df.iloc[0]
        df = df[1:]
        dfs.append(df)
    return dfs


class TimeDependentConnections:
    """
    Structure for reading/writing interactions

    A :class:`TimeDependentConnections` object is suitable when there are time dependent interactions between two quantities
    This class is used for both transfers and interactions. The content that it writes consists of

    - A connection matrix table that has Y/N selection of which interactions are present between two things
    - A set of pairwise connections specifying to, from, units, assumption, and time

    Interactions can have a diagonal, whereas transfers cannot (e.g. a population can infect itself but cannot transfer to itself).

    In Excel, a :class:`TimeDependentConnections` maps to three tables

    1. A table to enter the code name and full name
    2. An interactions matrix with Y/N indicating whether an interaction between two populations exists
    3. A set of rows for entering time varying data for each pair of populations

    :param code_name: the code name of this quantity e.g. 'aging'
    :param full_name: the full name of this quantity e.g. 'Aging'
    :param tvec: time values for the time-dependent rows
    :param pops: list of strings to use as the rows and columns - these are typically lists of population code names
    :param type: 'transfer' or 'interaction'. A transfer cannot have diagonal entries, and can have Number or Probability formats. An Interaction can have
                 diagonal entries and only has N.A. formats
    :param ts: Optionally specify a dict containing all of the non-empty TimeSeries objects used. The format is ``{(from_pop, to_pop):TimeSeries}``.
               An interaction can only be Y/N for clarity, if it is Y then a row is displayed for the TimeSeries. Actually, the Y/N can be
               decided in the first instance based on the provided TimeSeries i.e. if a TimeSeries is provided for an interaction, then the
               interaction must have been marked with Y
    :param pop_type: Specify pop_type, which is used by :meth:`ProjectData.add_pop` to determine which TDCs to add new populations to

    """

    def __init__(self, code_name: str, full_name: str, tvec: np.array, from_pops: list, to_pops: list, interpop_type: str, ts: dict = None, from_pop_type: str = None, to_pop_type: str = None):
        self.code_name = code_name
        self.full_name = full_name
        self.type = interpop_type
        self.from_pop_type = from_pop_type
        self.from_pops = from_pops
        self.to_pop_type = to_pop_type
        self.to_pops = to_pops
        self.tvec = tvec
        self.ts = ts if ts is not None else sc.odict()

        self.attributes = {}  #: Attributes associated with the table
        self.ts_attributes = {}  #: Attributes associated with each TimeSeries row
        self.ts_attributes["Provenance"] = {}  # Include provenance attribute by default

        self.assumption_heading = "Constant"  #: Heading to use for assumption column
        self.write_units = None  #: Write a column for units (if None, units will be written if any of the TimeSeries have units)
        self.write_uncertainty = None  #: Write a column for units (if None, units will be written if any of the TimeSeries have uncertainty)
        self.write_assumption = None  #: Write a column for units (if None, units will be written if any of the TimeSeries have an assumption)

        if self.type == "transfer":
            self.enable_diagonal = False
            self.allowed_units = []
            self.allowed_units.append("%s (%s)" % (FS.QUANTITY_TYPE_NUMBER.title(), format_duration(1, pluralize=True)))
            self.allowed_units.append("%s (per %s)" % (FS.QUANTITY_TYPE_RATE.title(), format_duration(1, pluralize=False)))
            self.allowed_units.append("%s (%s)" % (FS.QUANTITY_TYPE_DURATION.title(), format_duration(1, pluralize=True)))
        elif self.type == "interaction":
            self.enable_diagonal = True
            self.allowed_units = [FS.DEFAULT_SYMBOL_INAPPLICABLE]
        else:
            raise Exception('Unknown TimeDependentConnections type - must be "transfer" or "interaction"')

    def __repr__(self):
        return '<TDC %s "%s">' % (self.type.title(), self.code_name)

    @staticmethod
    def from_tables(tables: list, interaction_type):
        """
        Instantiate based on list of tables

        This method instantiates and initializes a new :class:`TimeDependentConnections` object from
        tables that have been read in using :func:`read_tables`. Note that the parent object
        such as :class:`ProjectData` is responsible for extracting the tables and passing them
        to this function. For instance, the transfers sheet might contain more than one set of
        tables, so it is the calling function's responsibility to split those tables up into
        the groups of three expected by this method.

        :param tables: A list of tables. A table here is a list of rows, and a row is a list of cells.
        :param interaction_type: A string identifying the interaction type - either 'transfer' or 'interaction'
        :return: A new :class:`TimeDependentConnections` instance

        """

        from .utils import TimeSeries  # Import here to avoid circular reference

        assert interaction_type in {"transfer", "interaction"}, "Unknown interaction type"

        # Read the TDC definition table (including attributes)
        code_name = None
        full_name = None
        from_pop_type = None
        to_pop_type = None

        attributes = {}
        for header_cell, value_cell in zip(tables[0][0], tables[0][1]):
            if header_cell.value is None:
                continue
            header = cell_get_string(header_cell)
            lowered_header = header.lower()
            if lowered_header == "abbreviation":
                code_name = cell_get_string(value_cell)
            elif lowered_header == "full name":
                full_name = cell_get_string(value_cell)
            elif lowered_header == "from population type":
                from_pop_type = cell_get_string(value_cell, True)
            elif lowered_header == "to population type":
                to_pop_type = cell_get_string(value_cell, True)
            else:
                attributes[header] = value_cell.value

        if interaction_type == "transfer":
            assert from_pop_type == to_pop_type, "Transfers can only occur between populations of the same type"

        if code_name is None:
            raise Exception("Code name/abbreviation missing")

        if full_name is None:
            raise Exception("Full name missing")

        # Read the pops from the Y/N table. The Y/N content of the table depends on the timeseries objects that
        # are present. That is, if the Y/N matrix contains a Y then a TimeSeries must be read in, and vice versa.
        # Therefore, we don't actually parse the matrix, and instead just read in all the TimeSeries instances
        # that are defined and infer the matrix that way.
        to_pops = [x.value for x in tables[1][0][1:] if x.value]
        from_pops = []
        for row in tables[1][1:]:
            from_pops.append(row[0].value)

        # Instantiate it
        tdc = TimeDependentConnections(code_name, full_name, None, from_pops=from_pops, to_pops=to_pops, interpop_type=interaction_type, from_pop_type=from_pop_type, to_pop_type=to_pop_type)
        tdc.attributes = attributes

        # Read the time series table
        headings = {}
        times = {}
        known_headings = {"from population", "to population", "units", "uncertainty", "constant", "assumption"}
        for i, cell in enumerate(tables[2][0]):
            v = cell.value
            if i == 0 or v is None:
                continue
            elif cell.data_type in {"s", "str"}:
                v = v.strip()
                if v.lower() in known_headings:
                    headings[v.lower()] = i
                else:
                    headings[v] = i
            elif cell.data_type == "n":
                if cell.is_date:
                    times[datetime_to_year(v)] = i
                else:
                    times[v] = i
            else:
                raise Exception("Unknown data type in cell %s of the spreadsheet - quantity must be a string or a number" % cell.coordinate)
        tdc.tvec = np.array(sorted(times), dtype=float)

        # Validate and process headings
        if not times and "constant" not in headings:
            raise Exception("Could not find an assumption or time-specific value - all tables must contain at least one of these values")
        tdc.write_units = True if "units" in headings else None
        tdc.write_uncertainty = True if "uncertainty" in headings else None
        tdc.write_assumption = True if "constant" in headings else None
        if "assumption" in headings:
            tdc.write_assumption = True
            tdc.assumption_heading = "Assumption"
        for heading in headings:
            if heading not in known_headings:
                # If it's not a known heading and it's a string, then it must be an attribute
                # Note that the way `headings` is populated by skipping i=0 ensures that the table name
                # is not interpreted as a heading
                tdc.ts_attributes[heading] = {}

        tdc.ts = sc.odict()
        for row in tables[2][1:]:
            if row[0].value != "...":
                assert row[0].value in from_pops, 'Population "%s" not found - should be contained in %s' % (row[0].value, from_pops)
                assert row[2].value in to_pops, 'Population "%s" not found - should be contained in %s' % (row[2].value, to_pops)
                vals = [x.value for x in row]
                from_pop = vals[0]
                to_pop = vals[2]

                if "units" in headings:
                    units = cell_get_string(row[headings["units"]], allow_empty=True)
                    if units.lower().strip() in FS.STANDARD_UNITS:
                        units = units.lower().strip()  # Only lower and strip units if they are standard units
                else:
                    units = None
                ts = TimeSeries(units=units)

                if "uncertainty" in headings:
                    ts.sigma = cell_get_number(row[headings["uncertainty"]])
                else:
                    ts.sigma = None

                if "constant" in headings:
                    ts.assumption = cell_get_number(row[headings["constant"]])
                elif "assumption" in headings:
                    ts.assumption = cell_get_number(row[headings["assumption"]])
                else:
                    ts.assumption = None

                for attribute in tdc.ts_attributes:
                    if attribute in headings:
                        tdc.ts_attributes[attribute][(from_pop, to_pop)] = row[headings[attribute]].value

                for t, idx in times.items():
                    ts.insert(t, cell_get_number(row[idx]))  # If cell_get_number returns None, this gets handled accordingly by ts.insert()

                tdc.ts[(from_pop, to_pop)] = ts

        return tdc

    def write(self, worksheet, start_row, formats, references: dict = None, widths: dict = None) -> int:
        """
        Write to cells in a worksheet

        :param worksheet: An xlsxwriter worksheet instance
        :param start_row: The first row in which to write values
        :param formats: Format dict for the opened workbook - typically the return value of :func:`standard_formats` when the workbook was opened
        :param references: References dict containing cell references for strings in the current workbook
        :param widths: ``dict`` storing column widths
        :return: The row index for the next available row for writing in the spreadsheet

        """

        assert self.assumption_heading in {"Constant", "Assumption"}, "Unsupported assumption heading"
        write_units = self.write_units if self.write_units is not None else any((ts.units is not None for ts in self.ts.values()))
        write_uncertainty = self.write_uncertainty if self.write_uncertainty is not None else any((ts.sigma is not None for ts in self.ts.values()))
        write_assumption = self.write_assumption if self.write_assumption is not None else any((ts.assumption is not None for ts in self.ts.values()))

        if not references:
            references = {x: x for x in self.from_pops + self.to_pops}  # Default null mapping for populations

        # First, write the name entry table
        current_row = start_row

        column = 0
        worksheet.write(current_row, column, "Abbreviation", formats["center_bold"])
        update_widths(widths, column, "Abbreviation")
        worksheet.write(current_row + 1, column, self.code_name)
        update_widths(widths, column, self.code_name)

        column += 1
        worksheet.write(current_row, column, "Full Name", formats["center_bold"])
        update_widths(widths, column, "Full Name")
        worksheet.write(current_row + 1, column, self.full_name)
        update_widths(widths, column, self.full_name)

        column += 1
        worksheet.write(current_row, column, "From population type", formats["center_bold"])
        update_widths(widths, column, "From population type")
        worksheet.write(current_row + 1, column, self.from_pop_type)
        update_widths(widths, column, self.from_pop_type)

        column += 1
        worksheet.write(current_row, column, "To population type", formats["center_bold"])
        update_widths(widths, column, "To population type")
        worksheet.write(current_row + 1, column, self.to_pop_type)
        update_widths(widths, column, self.to_pop_type)

        for attribute, value in self.attributes.items():
            column += 1
            worksheet.write(current_row, column, attribute, formats["center_bold"])
            update_widths(widths, column, attribute)
            worksheet.write(current_row + 1, column, value)
            update_widths(widths, column, value)

        references[self.code_name] = "='%s'!%s" % (worksheet.name, xlrc(current_row, 0, True, True))
        references[self.full_name] = "='%s'!%s" % (worksheet.name, xlrc(current_row, 1, True, True))  # Reference to the full name

        # Then, write the Y/N matrix
        current_row += 3  # Leave a blank row below the matrix
        # Note - table_references are local to this TimeDependentConnections instance
        # For example, there could be two transfers, and each of them could potentially transfer between 0-4 and 5-14
        # so the worksheet might contain two references from 0-4 to 5-14 but they would be for different transfers and thus
        # the time-dependent rows would depend on different boolean table cells
        current_row, table_references, values_written = self._write_pop_matrix(worksheet, current_row, formats, references, boolean_choice=True, widths=widths)

        # Finally, write the time dependent part
        headings = []
        headings.append("From population")
        headings.append("")  # --->
        headings.append("To population")
        offset = len(headings)

        attribute_index = {}
        for attribute in self.ts_attributes:
            attribute_index[attribute] = offset
            headings.append(attribute)
            offset += 1

        if write_units:
            headings.append("Units")
            units_index = offset  # Column to write the units in
            offset += 1

        if write_uncertainty:
            headings.append("Uncertainty")
            uncertainty_index = offset  # Column to write the units in
            offset += 1

        if write_assumption:
            headings.append(self.assumption_heading)
            headings.append("")
            constant_index = offset
            offset += 1
            if len(self.tvec):
                offset += 1  # Additional offset for the 'OR' column

        headings += [float(x) for x in self.tvec]
        for i, entry in enumerate(headings):
            if entry in references:
                worksheet.write_formula(current_row, 0, references[entry], formats["center_bold"], value=entry)
            else:
                worksheet.write(current_row, i, entry, formats["center_bold"])
            update_widths(widths, i, entry)

        # Now, we will write a wrapper that gates the content
        # If the gating cell is 'Y', then the content will be displayed, otherwise not
        def gate_content(content, gating_cell):
            if content.startswith("="):  # If this is itself a reference
                return '=IF(%s="Y",%s,"...")' % (gating_cell, content[1:])
            else:
                return '=IF(%s="Y","%s","...")' % (gating_cell, content)

        for from_idx in range(0, len(self.from_pops)):
            for to_idx in range(0, len(self.to_pops)):
                current_row += 1
                from_pop = self.from_pops[from_idx]
                to_pop = self.to_pops[to_idx]
                entry_tuple = (from_pop, to_pop)
                entry_cell = table_references[entry_tuple]

                # Write hyperlink
                if values_written[entry_cell] != FS.DEFAULT_SYMBOL_INAPPLICABLE:
                    worksheet.write_url(entry_cell, "internal:%s!%s" % (worksheet.name, xlrc(current_row, 1)), cell_format=formats["center_unlocked"], string=values_written[entry_cell])
                    worksheet.write_url(xlrc(current_row, 1), "internal:%s!%s" % (worksheet.name, entry_cell), cell_format=formats["center_unlocked"])

                if entry_tuple in self.ts:
                    ts = self.ts[entry_tuple]
                    format = formats["not_required"]
                else:
                    ts = None
                    format = formats["unlocked"]

                if ts:
                    worksheet.write_formula(current_row, 0, gate_content(references[from_pop], entry_cell), formats["center_bold"], value=from_pop)
                    update_widths(widths, 0, from_pop)
                    worksheet.write_formula(current_row, 1, gate_content("--->", entry_cell), formats["center"], value="--->")
                    worksheet.write_formula(current_row, 2, gate_content(references[to_pop], entry_cell), formats["center_bold"], value=to_pop)
                    update_widths(widths, 2, to_pop)

                    # Write the attributes
                    for attribute in self.ts_attributes:
                        if isinstance(self.ts_attributes[attribute], dict):
                            if entry_tuple in self.ts_attributes[attribute]:
                                val = self.ts_attributes[attribute][entry_tuple]
                            else:
                                val = None
                        else:
                            val = self.ts_attributes[attribute]

                        if val is not None:
                            worksheet.write(current_row, attribute_index[attribute], val)
                            update_widths(widths, attribute_index[attribute], val)

                    if self.write_units:
                        worksheet.write(current_row, units_index, ts.units, format)
                        update_widths(widths, units_index, ts.units)
                        if self.allowed_units:
                            worksheet.data_validation(xlrc(current_row, units_index), {"validate": "list", "source": [x for x in self.allowed_units]})

                    if self.write_uncertainty:
                        worksheet.write(current_row, uncertainty_index, ts.sigma, formats["not_required"])

                    if self.write_assumption:
                        worksheet.write(current_row, constant_index, ts.assumption, format)
                        if len(self.tvec):
                            worksheet.write_formula(current_row, constant_index + 1, gate_content("OR", entry_cell), formats["center"], value="OR")
                            update_widths(widths, constant_index + 1, "OR")

                else:
                    worksheet.write_formula(current_row, 0, gate_content(references[from_pop], entry_cell), formats["center_bold"], value="...")
                    worksheet.write_formula(current_row, 1, gate_content("--->", entry_cell), formats["center"], value="...")
                    worksheet.write_formula(current_row, 2, gate_content(references[to_pop], entry_cell), formats["center_bold"], value="...")

                    if self.write_units:
                        worksheet.write_blank(current_row, units_index, "", format)
                        if self.allowed_units:
                            worksheet.data_validation(xlrc(current_row, units_index), {"validate": "list", "source": [x for x in self.allowed_units]})

                    if self.write_uncertainty:
                        worksheet.write_blank(current_row, uncertainty_index, "", formats["not_required"])

                    if self.write_assumption:
                        worksheet.write_blank(current_row, constant_index, "", format)
                        if len(self.tvec):
                            worksheet.write_formula(current_row, constant_index + 1, gate_content("OR", entry_cell), formats["center"], value="...")
                            update_widths(widths, constant_index + 1, "...")

                content = [None] * len(self.tvec)

                if ts:
                    for t, v in zip(ts.t, ts.vals):
                        idx = np.where(self.tvec == t)[0][0]
                        content[idx] = v

                for idx, v in enumerate(content):
                    if v is None:
                        worksheet.write_blank(current_row, offset + idx, v, format)
                    else:
                        worksheet.write(current_row, offset + idx, v, format)
                    widths[offset + idx] = max(widths[offset + idx], 7) if offset + idx in widths else 7

                if not content:
                    idx = 0

                if self.write_assumption and len(self.tvec):
                    # Conditional formatting for the assumption, depending on whether time-values were entered
                    fcn_empty_times = 'COUNTIF(%s:%s,"<>" & "")>0' % (xlrc(current_row, offset), xlrc(current_row, offset + idx))
                    worksheet.conditional_format(xlrc(current_row, constant_index), {"type": "formula", "criteria": "=" + fcn_empty_times, "format": formats["ignored"]})
                    worksheet.conditional_format(xlrc(current_row, constant_index), {"type": "formula", "criteria": "=AND(%s,NOT(ISBLANK(%s)))" % (fcn_empty_times, xlrc(current_row, constant_index)), "format": formats["ignored_warning"]})

                # Conditional formatting for the row - it has a white background if the gating cell is 'N'
                worksheet.conditional_format("%s:%s" % (xlrc(current_row, 3), xlrc(current_row, offset + idx)), {"type": "formula", "criteria": '=%s<>"Y"' % (entry_cell), "format": formats["white_bg"]})

        current_row += 2

        return current_row

    def _write_pop_matrix(self, worksheet, start_row, formats, references: dict = None, boolean_choice=False, widths: dict = None):
        """
        Write a square matrix to Excel

        This function writes the Y/N matrix

        - Transfer matrix
        - Interactions matrix

        If ``self.enable_diagonal`` is ``False`` then the diagonal will be forced to be ``'N.A.'``. If an entry
        is specified for an entry on the diagonal and ``enable_diagonal=False``, an error will be thrown

        :param worksheet: An xlsxwriter worksheet instance
        :param start_row: The first row in which to write values
        :param formats: Format dict for the opened workbook - typically the return value of :func:`standard_formats` when the workbook was opened
        :param references: Optionally supply dict with references, used to link population names in Excel
        :param boolean_choice: If True, values will be coerced to Y/N and an Excel validation will be added
        :param widths: ``dict`` storing column widths
        :return: Tuple with ``(next_row, table_references, values_written)``. The references are used for hyperlinking to the Excel matrix

        """

        entries = self.ts

        if not references:
            references = {x: x for x in self.from_pops + self.to_pops}  # This is a null-mapping that takes say 'adults'->'adults' thus simplifying the workflow. Otherwise, it's assumed a reference exists for every node

        table_references = {}
        values_written = {}

        # Write the headers
        for i, node in enumerate(self.to_pops):
            worksheet.write_formula(start_row, i + 1, references[node], formats["center_bold"], value=node)
            update_widths(widths, i + 1, node)
        for i, node in enumerate(self.from_pops):
            worksheet.write_formula(start_row + i + 1, 0, references[node], formats["center_bold"], value=node)
            update_widths(widths, 0, node)

        # Prepare the content - first replace the dict with one keyed by index. This is because we cannot apply formatting
        # after writing content, so have to do the writing in a single pass over the entire matrix
        if boolean_choice:
            content = np.full((len(self.from_pops), len(self.to_pops)), "N", dtype=object)  # This will also coerce the value to string in preparation for writing
        else:
            content = np.full((len(self.from_pops), len(self.to_pops)), "", dtype=object)  # This will also coerce the value to string in preparation for writing

        for interaction, value in entries.items():
            from_pop, to_pop = interaction
            if not self.enable_diagonal and from_pop == to_pop:
                raise Exception("Trying to write a diagonal entry to a table that is not allowed to contain diagonal terms")  # This is because data loss will occur if the user adds entries on the diagonal, then writes the table, and then reads it back in
            from_idx = self.from_pops.index(from_pop)
            to_idx = self.to_pops.index(to_pop)
            if boolean_choice:
                value = "Y" if value else "N"
            content[from_idx, to_idx] = value

        # Write the content
        for from_idx in range(0, len(self.from_pops)):
            for to_idx in range(0, len(self.to_pops)):
                row = start_row + 1 + from_idx
                col = to_idx + 1
                if not self.enable_diagonal and self.to_pops[to_idx] == self.from_pops[from_idx]:  # Disable the diagonal if it's linking the same two quantities and that's desired
                    val = FS.DEFAULT_SYMBOL_INAPPLICABLE
                    worksheet.write(row, col, val, formats["center"])
                    worksheet.data_validation(xlrc(row, col), {"validate": "list", "source": ["N.A."]})
                else:
                    val = content[from_idx, to_idx]
                    worksheet.write(row, col, content[from_idx, to_idx], formats["center_unlocked"])
                    if boolean_choice:
                        worksheet.data_validation(xlrc(row, col), {"validate": "list", "source": ["Y", "N"]})
                        worksheet.conditional_format(xlrc(row, col), {"type": "cell", "criteria": "equal to", "value": '"Y"', "format": formats["unlocked_boolean_true"]})
                        worksheet.conditional_format(xlrc(row, col), {"type": "cell", "criteria": "equal to", "value": '"N"', "format": formats["unlocked_boolean_false"]})
                table_references[(self.from_pops[from_idx], self.to_pops[to_idx])] = xlrc(row, col, True, True)  # Store reference to this interaction
                values_written[table_references[(self.from_pops[from_idx], self.to_pops[to_idx])]] = val

        next_row = start_row + 1 + len(self.from_pops) + 1
        return next_row, table_references, values_written


class TimeDependentValuesEntry:
    """Table for time-dependent data entry

    This class is Databooks and Program books to enter potentially time-varying data.
    Conceptually, it maps a set of TimeSeries object to a single name and table in the
    spreadsheet. For example, a Characteristic might contain a TimeSeries for each population,
    and the resulting TimeDependentValuesEntry (TDVE) table would have a `name` matching the
    population, and TimeSeries for each population.

    The TDVE class optionally allows the specification of units, assumptions, and uncertainty,
    which each map to properties on the underlying TimeSeries objects. It also contains a
    time vector corresponding to the time values that appear or will appear in the spreadsheet.

    Note that the units are stored within the TimeSeries objects, which means that they can
    are able to differ across rows.

    :param name: The name/title for this table
    :param tvec: Specify the time values for this table. All TimeSeries in the ts dict should have corresponding time values
    :param ts: Optionally specify an odict() of TimeSeries objects populating the rows. Could be populated after
    :param allowed_units: Optionally specify a list of allowed units that will appear as a dropdown
    :param comment: Optionally specify descriptive text that will be added as a comment to the name cell

    """

    def __init__(self, name, tvec: np.array = None, ts=None, allowed_units: list = None, comment: str = None):

        if ts is None:
            ts = sc.odict()

        self.name = name  #: Name for th quantity printed in Excel
        self.comment = comment  #: A comment that will be added in Excel
        self.tvec = [] if tvec is None else tvec  #: time axis (e.g. np.arange(2000,2019)) - all TimeSeries time values must exactly match one of the values here
        self.ts = ts  # : dict of :class:`TimeSeries` objects
        self.allowed_units = [x.title() if x in FS.STANDARD_UNITS else x for x in allowed_units] if allowed_units is not None else None  # Otherwise, can be an odict with keys corresponding to ts - leave as None for no restriction

        self.ts_attributes = {}  #: Dictionary containing extra attributes to write along with each TimeSeries object.
        self.ts_attributes["Provenance"] = {}  # Include provenance attribute by default

        #  Keys are attribute name, values can be either a scalar or a dict keyed by the same keys as self.ts. Compared to units, uncertainty etc.
        #  attributes are store in the TDVE rather than in the TimeSeries
        self.assumption_heading = "Constant"  #: Heading to use for assumption column

        self.write_units = None  #: Write a column for units (if None, units will be written if any of the TimeSeries have units)
        self.write_uncertainty = None  #: Write a column for units (if None, units will be written if any of the TimeSeries have uncertainty)
        self.write_assumption = None  #: Write a column for units (if None, units will be written if any of the TimeSeries have an assumption)

    def __repr__(self):
        output = sc.prepr(self)
        return output

    @property
    def has_data(self) -> bool:
        """
        Check whether all time series have data entered

        :return: True if all of the TimeSeries objects stored in the TDVE have data
        """

        return all([x.has_data for x in self.ts.values()])

    @staticmethod
    def from_rows(rows: list):
        """
        Create new instance from Excel rows

        Given a set of openpyxl rows, instantiate a :class:`TimeDependentValuesEntry` object
        That is, the parent object e.g. :class:`ProjectData` is responsible for finding where the TDVE table is,
        and reading all of the rows associated with it (skipping ``#ignored`` rows) and then passing those rows,
        unparsed, to this function

        Headings for 'units', 'uncertainty', and 'assumption'/'constant' are optional and will be read in
        if they are present in the spreadsheet.

        :param rows: A list of rows
        :return: A new :class:`TimeDependentValuesEntry` instance

        """

        from .utils import TimeSeries  # Import here to avoid circular reference

        # Retrieve the name
        name = rows[0][0].value
        if name is None:
            raise Exception('The name of the table is missing. This can also happen if extra rows have been added without a "#ignore" entry in the first column')
        elif not sc.isstring(name):
            raise Exception("In cell %s of the spreadsheet, the name of the quantity assigned to this table needs to be a string" % rows[0][0].coordinate)
        name = name.strip()  # The name needs to be written back in a case sensitive form

        tdve = TimeDependentValuesEntry(name)

        # Read the headings
        headings = {}
        times = {}
        known_headings = {"units", "uncertainty", "constant", "assumption"}
        for i, cell in enumerate(rows[0]):
            v = cell.value
            if i == 0 or v is None:
                continue
            elif cell.data_type in {"s", "str"}:
                v = v.strip()
                if v.lower() in known_headings:
                    headings[v.lower()] = i
                else:
                    headings[v] = i
            elif cell.is_date:
                times[datetime_to_year(v)] = i
            elif cell.data_type == "n":
                times[v] = i
            else:
                raise Exception("Unknown data type in cell %s of the spreadsheet - quantity must be a string or a number" % cell.coordinate)
        tdve.tvec = np.array(sorted(times), dtype=float)

        # Validate and process headings
        if not times and "constant" not in headings:
            raise Exception("Could not find an assumption or time-specific value - all tables must contain at least one of these values")
        tdve.write_units = True if "units" in headings else None
        tdve.write_uncertainty = True if "uncertainty" in headings else None
        tdve.write_assumption = True if "constant" in headings else None
        if "assumption" in headings:
            tdve.write_assumption = True
            tdve.assumption_heading = "Assumption"
        for heading in headings:
            if heading not in known_headings:
                # If it's not a known heading and it's a string, then it must be an attribute
                # Note that the way `headings` is populated by skipping i=0 ensures that the table name
                # is not interpreted as a heading
                tdve.ts_attributes[heading] = {}
        ts_entries = sc.odict()

        for row in rows[1:]:
            if not row[0].data_type in {"s", "str"}:
                raise Exception("In cell %s of the spreadsheet, the name of the entry was expected to be a string, but it was not. The left-most column is expected to be a name. If you are certain the value is correct, add an single quote character at the start of the cell to ensure it remains as text" % row[0].coordinate)
            series_name = row[0].value.strip()

            if "units" in headings:
                units = cell_get_string(row[headings["units"]], allow_empty=True)
                if units.lower().strip() in FS.STANDARD_UNITS:
                    units = units.lower().strip()  # Only lower and strip units if they are standard units
            else:
                units = None
            ts = TimeSeries(units=units)

            if "uncertainty" in headings:
                ts.sigma = cell_get_number(row[headings["uncertainty"]])
            else:
                ts.sigma = None

            if "constant" in headings:
                ts.assumption = cell_get_number(row[headings["constant"]])
            elif "assumption" in headings:
                ts.assumption = cell_get_number(row[headings["assumption"]])
            else:
                ts.assumption = None

            for attribute in tdve.ts_attributes:
                if attribute in headings:
                    # If it's a default attribute e.g. provenance, and it is missing from the databook, then don't populate it
                    tdve.ts_attributes[attribute][series_name] = row[headings[attribute]].value

            for t, idx in times.items():
                ts.insert(t, cell_get_number(row[idx]))  # If cell_get_number returns None, this gets handled accordingly by ts.insert()
            ts_entries[series_name] = ts

        tdve.ts = ts_entries
        return tdve

    def write(self, worksheet, start_row, formats, references: dict = None, widths: dict = None) -> int:
        """
        Write to cells in a worksheet

        Note that the year columns are drawn from the ``tvec`` attribute. To suppress the year columns (e.g. for the user to enter only an assumption)
        then set ``tvec`` to an empty array/list.

        :param worksheet: An xlsxwriter worksheet instance
        :param start_row: The first row in which to write values
        :param formats: Format dict for the opened workbook - typically the return value of :func:`standard_formats` when the workbook was opened
        :param references: References dict containing cell references for strings in the current workbook
        :param widths: ``dict`` storing column widths
        :return: The row index for the next available row for writing in the spreadsheet

        """

        assert self.assumption_heading in {"Constant", "Assumption"}, "Unsupported assumption heading"

        write_units = self.write_units if self.write_units is not None else any((ts.units is not None for ts in self.ts.values()))
        write_uncertainty = self.write_uncertainty if self.write_uncertainty is not None else any((ts.sigma is not None for ts in self.ts.values()))
        write_assumption = self.write_assumption if self.write_assumption is not None else any((ts.assumption is not None for ts in self.ts.values()))

        if not references:
            references = dict()

        current_row = start_row

        # First, assemble and write the headings
        headings = []
        headings.append(self.name)
        offset = 1  # This is the column where the time values start (after the 'or')

        # Next allocate attributes
        attribute_index = {}
        for attribute in self.ts_attributes:
            attribute_index[attribute] = offset
            headings.append(attribute)
            offset += 1

        if write_units:
            headings.append("Units")
            units_index = offset  # Column to write the units in
            offset += 1

        if write_uncertainty:
            headings.append("Uncertainty")
            uncertainty_index = offset  # Column to write the units in
            offset += 1

        if write_assumption:
            headings.append(self.assumption_heading)
            headings.append("")
            constant_index = offset
            offset += 2

        headings += [float(x) for x in self.tvec]
        for i, entry in enumerate(headings):
            if entry in references:
                worksheet.write_formula(current_row, 0, references[entry], formats["center_bold"], value=entry)
            else:
                worksheet.write(current_row, i, entry, formats["center_bold"])
            update_widths(widths, i, entry)

        if self.comment:
            worksheet.write_comment(xlrc(current_row, 0), self.comment)

        # Now, write the TimeSeries objects - self.ts is an odict and whatever pops are present will be written in whatever order they are in
        for row_name, row_ts in self.ts.items():
            current_row += 1

            # Write the name
            if row_name in references:
                worksheet.write_formula(current_row, 0, references[row_name], formats["center_bold"], value=row_name)
                update_widths(widths, 0, row_name)
            else:
                worksheet.write_string(current_row, 0, row_name, formats["center_bold"])
                update_widths(widths, 0, row_name)

            # Write the attributes
            for attribute in self.ts_attributes:
                if isinstance(self.ts_attributes[attribute], dict):
                    if row_name in self.ts_attributes[attribute]:
                        val = self.ts_attributes[attribute][row_name]
                    else:
                        val = None
                else:
                    val = self.ts_attributes[attribute]

                if val is not None:
                    worksheet.write(current_row, attribute_index[attribute], val)
                    update_widths(widths, attribute_index[attribute], val)

            # Write the units
            if write_units:
                if row_ts.units:
                    if row_ts.units.lower().strip() in FS.STANDARD_UNITS:  # Preserve case if nonstandard unit
                        unit = row_ts.units.title().strip()
                    else:
                        unit = row_ts.units.strip()
                    worksheet.write(current_row, units_index, unit)
                    update_widths(widths, units_index, unit)
                else:
                    worksheet.write(current_row, units_index, FS.DEFAULT_SYMBOL_INAPPLICABLE)

                if self.allowed_units and isinstance(self.allowed_units, dict) and row_name in self.allowed_units:  # Add dropdown selection if there is more than one valid choice for the units
                    allowed = self.allowed_units[row_name]
                elif self.allowed_units and not isinstance(self.allowed_units, dict):
                    allowed = self.allowed_units
                else:
                    allowed = None

                if allowed:
                    worksheet.data_validation(xlrc(current_row, units_index), {"validate": "list", "source": allowed})

            if write_uncertainty:
                if row_ts.sigma is None:
                    worksheet.write(current_row, uncertainty_index, row_ts.sigma, formats["not_required"])  # NB. For now, uncertainty is always optional
                else:
                    worksheet.write(current_row, uncertainty_index, row_ts.sigma, formats["not_required"])

            if row_ts.has_data:
                format = formats["not_required"]
            else:
                format = formats["unlocked"]

            if write_assumption:
                worksheet.write(current_row, constant_index, row_ts.assumption, format)
                if len(self.tvec):
                    worksheet.write(current_row, constant_index + 1, "OR", formats["center"])
                    update_widths(widths, constant_index + 1, "OR")

            # Write the time values if they are present
            if len(self.tvec):
                content = [None] * len(self.tvec)  # Initialize an empty entry for every time in the TDVE's tvec

                for t, v in zip(row_ts.t, row_ts.vals):
                    # If the TimeSeries contains data for that time point, then insert it now
                    idx = np.where(self.tvec == t)[0]
                    if len(idx):
                        content[idx[0]] = v

                for idx, v in enumerate(content):
                    if v is None:
                        worksheet.write_blank(current_row, offset + idx, v, format)
                    else:
                        worksheet.write(current_row, offset + idx, v, format)
                    widths[offset + idx] = max(widths[offset + idx], 7) if offset + idx in widths else 7

                if write_assumption:
                    # Conditional formatting for the assumption
                    # Do this here, because after the loop above, we have easy and clear access to the range of cells to include in the formula
                    fcn_empty_times = 'COUNTIF(%s:%s,"<>" & "")>0' % (xlrc(current_row, offset), xlrc(current_row, offset + len(content) - 1))
                    # Hatched out if the cell will be ignored
                    worksheet.conditional_format(xlrc(current_row, constant_index), {"type": "formula", "criteria": "=" + fcn_empty_times, "format": formats["ignored"]})
                    worksheet.conditional_format(xlrc(current_row, constant_index), {"type": "formula", "criteria": "=AND(%s,NOT(ISBLANK(%s)))" % (fcn_empty_times, xlrc(current_row, constant_index)), "format": formats["ignored_warning"]})

        return current_row + 2  # Add two so there is a blank line after this table


def cell_get_string(cell, allow_empty=False) -> str:
    """
    Return string value from cell

    This function checks if a cell contains a string. If it does, the stripped value
    will be returned. Otherwise, an informative error will be raised

    Note that the string type is determined from the cell's value rather than
    the openpyxl cell data type.

    :param cell: An openpyxl cell
    :return: A string with the contents of the cell
    """
    if cell.value is None and allow_empty:
        return None
    elif not sc.isstring(cell.value):
        raise Exception("Cell %s needs to contain a string (i.e. not a number, date, or other cell type)" % cell.coordinate)
    else:
        return cell.value.strip()


def cell_get_number(cell, dtype=float):
    """
    Return numeric value from cell

    This function is to guard against accidentally having the Excel cell contain a string
    instead of a number. If a string has been entered, an error will be raised. The added value
    from this function is that if the Excel cell type is empty but the value is
    empty or ``N.A.`` then the value will be treated as though the cell was correctly set to a
    numeric type but had been left empty.

    The output is cast to ``dtype`` which means that code that requires numeric input from Excel
    can use this input to guarantee that the resulting number is of the correct type, or ``None``.

    :param cell: An openpyxl cell
    :param dtype: If the cell is numeric, cast to this type (default is `float` but could be `int` for example)
    :return: A scalar instance of ``dtype`` (e.g. ``float``) or ``None`` if cell is empty or being treated as empty
    :raises: :class:`Exception` if the cell contains a string

    """

    if cell.value is None:
        return None
    elif cell.data_type == "n":  # Numeric type
        return dtype(cell.value)
    elif cell.data_type == "s":  # Only do relatively expensive string processing if it's actually a string type
        s = cell.value.lower().strip()
        if s == FS.DEFAULT_SYMBOL_INAPPLICABLE:
            return None
        elif not s.replace("-", ""):
            return None

    raise Exception("Cell %s needs to contain a number" % cell.coordinate)


def validate_category(workbook, expected_category) -> None:
    """
    Check Atomica workbook type

    This function makes sure that a workbook has a particular category property
    stored within it, and displays an appropriate error message if not. If the
    category isn't present or doesn't start with 'atomica', just ignore it for
    robustness (instead, a parsing error will likely be raised)

    :param workbook: An openpyxl workbook
    :param category: The expected string category
    :raises: :class:`Exception` if the workbook category is not valid

    """

    category = workbook.properties.category
    if category and sc.isstring(category) and category.startswith("atomica:"):
        if category.strip() != expected_category.strip():
            expected_type = expected_category.split(":")[1].title()
            actual_type = category.split(":")[1].title()
            message = "Error loading %s - the provided file was a %s file" % (expected_type, actual_type)
            raise Exception(message)
