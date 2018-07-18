import os
import numpy as np

import xlrd
import sciris.core as sc
from xlsxwriter.utility import xl_rowcol_to_cell as xlrc
from .system import AtomicaException
import openpyxl

# This file reads and writes Tables from a spreadsheet.

class AtomicaWorkbook(object):
    # This is the intermediate representation of a workbook

    def __init__(self,fname=None):
        self.sheets = sc.odict()

        if fname:
            self.read(fname)

    def read(self, fname):
        self.sheets = sc.odict()

        wb = openpyxl.load_workbook(filename=fname, data_only=True, read_only = False)

        for ws in wb.worksheets:

            tables = [] # Tables in this sheet

            table_rows = [] # Rows to assign to the next table
            for row in ws.rows:

                if row[0].value and row[0].value.startswith('#ignore'):
                    continue # Move on to the next row if row skipping is marked True

                for cell in row:
                    if cell.value:
                        table_rows.append(row)
                        break
                else:
                    if table_rows:
                        tables.append(Table.from_rows(table_rows))
                        table_rows = []
            if table_rows: # Flush the final row
                tables.append(Table.from_rows(table_rows))

            self.sheets[ws.title] = tables

        return

    def write(self,fname):
        wb = openpyxl.Workbook(write_only = True)

        for name, tables in self.sheets.items():
            ws = wb.create_sheet(name)
            for table in tables:
                table.write(ws)

        wb.save(fname)

        return

class Table(object):
    # A Table is the fundamental parsing unit - it is a rectangular set of cells together with (optional) formatting and comments
    # Some particular Tables could have programatically dependent contents e.g. TDVE tables have logic for gating the assumption value
    # But the base table is just raw IO

    def __init__(self):
        # By convention, the first row of content are the headings for this table
        # Could extend the members to include other features e.g. bold cells
        self.content = [] # A nested list for the rectangular Table object
        self.comments = []
        self.fill = []

    # Use the static method below to instantiate a Table when parsing an Excel file
    @staticmethod
    def from_rows(rows):
        # The first row defines headings
        self = Table()

        for i in range(len(rows[0]),0,-1):
            if rows[0][i-1].value:
                extent = i
                break
        else:
            raise AtomicaException('Heading row was empty')

        for row in rows:
            row_content = []
            row_comments = []
            row_fill = []

            for i in range(0,extent):
                cell = row[i]

                # If we are being constructed using an array of values rather than an array of Cells
                if not isinstance(cell,openpyxl.cell.cell.Cell):
                    row_content.append(cell)
                    row_comments.append(None)
                    row_fill.append(None)
                    continue

                row_content.append(cell.value)

                if cell.comment:
                    row_comments.append(cell.comment.content)
                else:
                    row_comments.append(None)

                if cell.fill.patternType == 'solid':
                    row_fill.append(cell.fill.fgColor.rgb)
                else:
                    row_fill.append(None)

            self.content.append(row_content)
            self.comments.append(row_comments)
            self.fill.append(row_fill)

        return self

    def write(self,ws):
        # Append this table to the given worksheet
        first_row = True
        for row_content,row_comments,row_fill in zip(self.content,self.comments,self.fill):
            row = []
            for content,comments,fill in zip(row_content,row_comments,row_fill):
                if not content:
                    row.append(None)
                else:
                    cell = openpyxl.worksheet.write_only.WriteOnlyCell(ws,value=content) # Indexing in Excel is 1-based (e.g. cell A1)

                    if comments:
                        cell.comment = openpyxl.comments.Comment(comments,author=None)
                    if fill:
                        cell.fill = openpyxl.styles.PatternFill(fgColor=fill, fill_type = "solid") # Solid PatternFill uses fgColor property for background

                    if first_row:
                        cell.font = openpyxl.styles.Font(bold=True)

                    row.append(cell)
            first_row = False
            ws.append(row)
        ws.append([None]*len(self.content[0]))





if __name__ == '__main__':
    b = AtomicaWorkbook('E:\\projects\\atomica\\atomica\\tests\\databooks\\databook_tb.xlsx')
    b.write('E:\\projects\\atomica\\atomica\\tests\\databooks\\openpyxl.xlsx')


