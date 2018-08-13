# -*- coding: utf-8 -*-
"""
Atomica data file.
Sets out a structure to store context-specific databook-imported values relating to a model.
"""

from .structure import TimeSeries
import sciris.core as sc
from xlsxwriter.utility import xl_rowcol_to_cell as xlrc
import openpyxl
from .excel import standard_formats, AtomicaSpreadsheet, read_tables, TimeDependentValuesEntry, TimeDependentConnections, apply_widths, update_widths, ProgramEntry
import xlrd
import xlsxwriter as xw
import io
import numpy as np
from .system import AtomicaException
from .structure import FrameworkSettings as FS
from collections import defaultdict
from six import string_types

# Data maps to a databook
# On construction, we first make some blank data, and then we write a databook in the same way as if we actually had
# data values
class ProjectData(object):

    def __init__(self):
        # This is just an overview of the structure of ProjectData
        # There are two pathways to a ProjectData
        # - Could load an existing one, with ProjectData.from_spreadsheet()
        # - Could make a new one, with ProjectData.new()
        self.pops = sc.odict() # This is an odict mapping code_name:{'label':full_name}
        self.transfers = list()
        self.interpops = list()
        self.tvec = None # This is the data's tvec used when instantiating new tables. Not _guaranteed_ to be the same for every TDVE/TDC table
        self.tdve = {}
        self.tdve_pages = sc.odict()

        # Internal storage used with methods while writing
        self._formats = None
        self._book = None
        self._references = None
    
    def __repr__(self):
        output = sc.desc(self)
        return output

    @property
    def start_year(self):
        # The ProjectData start year is defined as the earliest time point in
        # any of the TDVE/TDC tables. This should be used when changing the simulation
        # start year
        start_year = np.inf
        for td_table in list(self.tdve.values()) + self.transfers + self.interpops:
            start_year = min(start_year,np.amin(td_table.tvec))
        return start_year

    @property
    def end_year(self):
        # The ProjectData start year is defined as the earliest time point in
        # any of the TDVE/TDC tables. This should be used when changing the simulation
        # start year
        end_year = -np.inf
        for td_table in list(self.tdve.values()) + self.transfers + self.interpops:
            end_year = max(end_year,np.amax(td_table.tvec))
        return end_year

    def change_tvec(self, tvec):
        # Set tvec in all TDVE/TDC tables contained in the ProjectData
        # - Setting `ProjectData.tvec = <>` will only affect tables created afterwards
        # - Calling `ProjectData.change_tvec()` will modify all existing tables
        self.tvec = tvec.copy()
        for td_table in list(self.tdve.values()) + self.transfers + self.interpops:
            td_table.tvec = tvec

    def get_ts(self,name,key=None):
        # This extracts a TimeSeries from a TDVE table or TDC table
        #
        # INPUTS
        # - name: - The code name of a transfer, interaction, or compartment/characteristic/parameter
        #         - The name of a transfer parameter instantiated in model.build e.g. 'age_0-4_to_5-14'.
        #           this is mainly useful when retrieving data for plotting, where variables are organized according
        #           to names like 'age_0-4_to_5-14'
        # - key: - If `name` is a comp/charac/par, then key should be a pop name
        #        - If `name` is a transfer or interaction, then key should be a tuple (from_pop,to_pop)
        #        - If `name` is the name of a model transfer parameter, then `key` should be left as `None`
        #
        # If the key was not found, return None

        # First, check if it's the name of a TDVE
        if name in self.tdve:
            if key in self.tdve[name].ts:
                return self.tdve[name].ts[key]

        # Then, if the key is None, we are working on a transfer parameter. So reconstruct the key
        if key is None:
            x = name.split('_to_')
            code_name,from_pop = x[0].split('_',1)
            to_pop = x[1]
            name = code_name
            key = (from_pop,to_pop)

        for tdc in self.transfers + self.interpops:
            if name == tdc.code_name:
                return tdc.ts[key]

        return None

    @staticmethod
    def new(framework,tvec,pops,transfers):
        # Make a brand new databook
        # pops - Can be a number, or an odict with names and labels
        # transfers - Can be a number, or an odict with names and labels
        # interactions - Can be a number, or an odict with names and labels

        if sc.isnumber(pops):
            new_pops = sc.odict()
            for i in range(0,pops):
                new_pops['pop_%d' % (i)] = 'Population %d' % (i)
        else:
            new_pops = pops

        if not new_pops:
            raise AtomicaException('A new databook must have at least 1 population')

        if sc.isnumber(transfers):
            new_transfers = sc.odict()
            for i in range(0,transfers):
                new_transfers['transfer_%d' % (i)] = 'Transfer %d' % (i)
        else:
            new_transfers = transfers

        # Make all of the empty TDVE objects - need to store them by page, and the page information is in the Framework
        data = ProjectData()
        data.tvec = tvec
        pages = defaultdict(list) # This will store {sheet_name:(code_name,databook_order)} which will then get sorted further

        for df in [framework.comps, framework.characs, framework.pars]:
            for _,spec in df.iterrows():
                databook_page = spec.get('databook page')
                databook_order = spec.get('databook order')
                full_name = spec['display name']
                if databook_page is not None:
                    if databook_order is None:
                        order = np.inf
                    else:
                        order = databook_order
                    pages[databook_page].append((full_name,order))
                    data.tdve[full_name] = TimeDependentValuesEntry(name=full_name,tvec=tvec,allowed_units=framework.get_allowed_units(full_name))

        # Now convert pages to full names and sort them into the correct order
        for _,spec in framework.sheets['databook pages'][0].iterrows():

            if spec['datasheet code name'] in pages:
                pages[spec['datasheet code name']].sort(key=lambda x: x[1])
                data.tdve_pages[spec['datasheet title']] = [x[0] for x in pages[spec['datasheet code name']]]
            else:
                data.tdve_pages[spec['datasheet title']] = list()

        # Now, proceed to add pops, transfers, and interactions
        for code_name,full_name in new_pops.items():
            data.add_pop(code_name,full_name)

        for code_name,full_name in new_transfers.items():
            data.add_transfer(code_name,full_name)

        for _,spec in framework.interactions.iterrows():
            interpop = data.add_interaction(spec.name, spec['display name'])
            if 'default_value' in spec and spec['default_value']:
                for from_pop in interpop.pops:
                    for to_pop in interpop.pops:
                        ts = TimeSeries(format=interpop.allowed_units[0],units=interpop.allowed_units[0])
                        ts.insert(None,spec['default_value'])
                        interpop.ts[(from_pop,to_pop)] = ts

        # Finally, insert parameter and characteristic default values
        for df in [framework.comps,framework.characs, framework.pars]:
            for _,spec in df.iterrows():
                # In order to write a default value
                # - The default value should be present and not None
                # - The quantity should appear in the databook
                if 'default value' in spec and (spec['default value'] is not None) and spec['databook page']:
                    tdve = data.tdve[spec['display name']]
                    for ts in tdve.ts.values():
                        ts.insert(None,spec['default value'])

        return data

    @staticmethod
    def from_spreadsheet(spreadsheet,framework):
        # The framework is needed because ProjectData
        # Instantiate a new Databook given a spreadsheet

        # Basically the strategy is going to be
        # 1. Read in all of the stuff - pops, transfers, interpops can be directly added to Data
        # 2. Read in all the other TDVE content, and then store it in the data specs according to the variable type defined in the Framework
        # e.g. the fact that 'Alive' is a Characteristic is stored in the Framework and Data but not in the Databook. So for example, we read in
        # a TDVE table called 'Alive', but it needs to be stored in data.specs['charac']['ch_alive'] and the 'charac' and 'ch_alive' are only available in the Framework
        #
        # spreadsheet - A ScirisSpreadsheet object
        # framework - A ProjectFramework object
        #
        # This static method will return a new Databook instance given the provided databook Excel file and Framework
        self = ProjectData()

        if isinstance(spreadsheet,string_types):
            spreadsheet = AtomicaSpreadsheet(spreadsheet)

        workbook = openpyxl.load_workbook(spreadsheet.get_file(),read_only=True,data_only=True) # Load in read-only mode for performance, since we don't parse comments etc.

        # These sheets are optional - if none of these are provided in the databook
        # then they will remain empty
        self.transfers = list()
        self.interpops = list()

        for sheet in workbook.worksheets:
            if sheet.title.startswith('#ignore'):
                continue

            if sheet.title == 'Population Definitions':
                self._read_pops(sheet)
            elif sheet.title == 'Transfers':
                self._read_transfers(sheet)
            elif sheet.title == 'Interactions':
                self._read_interpops(sheet)
            elif sheet.title == 'Metadata':
                continue
            else:
                self.tdve_pages[sheet.title] = []
                for table in read_tables(sheet):
                    tdve = TimeDependentValuesEntry.from_rows(table)

                    # If this fails, the TDVE was not found in the framework. That's a critical stop error, because the framework needs to at least declare what kind of variable this is
                    code_name = framework.get_variable(tdve.name)[0].name
                    tdve.allowed_units = [x.title() for x in framework.get_allowed_units(code_name)]

                    self.tdve[code_name] = tdve
                    # Store the TDVE on the page it was actually on, rather than the one in the framework. Then, if users move anything around, the change will persist
                    self.tdve_pages[sheet.title].append(code_name)

        # Set the ProjectData's tvec based on the first TDVE table
        # 99.9% of the time, all of the tables will have the same values and so this is generally safe
        # The only time unexpected behaviour might occur is if the first TDVE table has exotic data points
        # and the user loads the databook, then adds a new transfer/interpop, the new table will have those same
        # modified data points. But what does the user expect, if the databook has mixed times
        self.tvec = self.tdve[self.tdve_pages[0][0]].tvec

        return self

    def validate(self,framework):
        # Check if the contents of the ProjectData can be used to run simulations
        #
        # A databook can be 'valid' in two senses
        #
        # - The Excel file adheres to the correct syntax and it can be parsed into a ProjectData object
        # - The resulting ProjectData object contains sufficient information to run a simulation
        #
        # Sometimes it is desirable for ProjectData to be valid in one sense rather than the other. For example,
        # in order to run a simulation, the ProjectData needs to contain at least one value for every TDVE table.
        # However, the TDVE table does _not_ need to contain values if all we want to do is add another key pop
        # Thus, the first stage of validation is the ProjectData constructor - if that runs, then users can
        # access methods like 'add_pop','remove_transfer' etc.
        #
        # However, to actually run a simulation, the _contents_ of the databook need to satisfy various conditions
        # These tests are implemented here. The typical workflow would be that ProjectData.validate() should be used
        # if a simulation is going to be run. In the first instance, this can be done in `Project.load_databook` but
        # the FE might want to perform this check at a different point if the databook manipulation methods e.g.
        # `add_pop` are going to be exposed in the interface
        #
        # This function throws an informative error if there are any problems identified or otherwise returns True
        #
        # Make sure that all of the quantities the Framework says we should read in have been read in, and that
        # those quantities all have some data values associated with them
        for df in [framework.comps, framework.characs, framework.pars]:
            for _, spec in df.iterrows():
                if spec['databook page'] is not None:
                    if spec.name not in self.tdve:
                        raise AtomicaException('Databook did not find any values for "%s" (%s)' % (spec['display name'],spec.name))
                    else:
                        allowed_units = framework.get_allowed_units(spec.name)
                        for name,ts in self.tdve[spec.name].ts.items():
                            assert name in self.pops, 'Population "%s" in "%s" not recognized. Should be one of: %s' % (name,self.tdve[spec.name].name,self.pops.keys())
                            assert ts.has_data, 'Data values missing for %s (%s)' % (self.tdve[spec.name].name, name)
                            assert ts.format is not None, 'Formats missing for %s (%s)' % (self.tdve[spec.name].name, name)
                            assert ts.units is not None, 'Units missing for %s (%s)' % (self.tdve[spec.name].name, name)
                            if allowed_units:
                                assert ts.units in allowed_units, 'Unit "%s" for %s (%s) do not match allowed units (%s)' % (ts.units,self.tdve[spec.name].name, name,allowed_units)

        for _,spec in framework.interactions.iterrows():
            for tdc in self.interpops:
                if tdc.code_name == spec.name:
                    for (to_pop,from_pop),ts in tdc.ts.items():
                        assert to_pop in self.pops, 'Population "%s" in "%s" not recognized. Should be one of: %s' % (name, self.tdve[spec.name].name, self.pops.keys())
                        assert from_pop in self.pops, 'Population "%s" in "%s" not recognized. Should be one of: %s' % (name, self.tdve[spec.name].name, self.pops.keys())
                        assert ts.has_data, 'Data values missing for interaction %s, %s->%s' % (spec.name, to_pop,from_pop)
                        assert ts.units.strip().title() == FS.DEFAULT_SYMBOL_INAPPLICABLE.title()
                    break
            else:
                raise AtomicaException('Required interaction "%s" not found in databook' % spec.name)

        return True

    def to_spreadsheet(self):
        # Initialize the bytestream
        f = io.BytesIO()

        # Open a workbook
        self._book = xw.Workbook(f)
        self._formats = standard_formats(self._book)
        self._references = {} # Reset the references dict

        # Write the contents
        self._write_pops()
        self._write_tdve()
        self._write_interpops()
        self._write_transfers()

        # Close the workbook
        self._book.close()

        # Dump the file content into a ScirisSpreadsheet
        spreadsheet = AtomicaSpreadsheet(f)

        # Clear everything
        f.close()
        self._book = None
        self._formats = None
        self._references = None

        # Return the spreadsheet
        return spreadsheet

    def save(self,fname):
        # Shortcut for saving to disk - FE RPC will probably use `to_spreadsheet()` but BE users will probably use `save()`
        ss = self.to_spreadsheet()
        ss.save(fname)

    def add_pop(self,code_name,full_name):
        # Add a population with the given name and label (full name)
        assert code_name not in self.pops, 'Population with name "%s" already exists' % (code_name)

        if code_name.strip().lower() == 'all':
            raise AtomicaException('A population was named "all", which is a reserved keyword and cannot be used as a population name')

        self.pops[code_name] = {'label':full_name}
        for interaction in self.transfers+self.interpops:
            interaction.pops.append(code_name)
        for tdve in self.tdve.values():
            default_unit = tdve.allowed_units[0] if tdve.allowed_units else None
            tdve.ts[code_name] = TimeSeries(format=default_unit,units=default_unit)

    def remove_pop(self,pop_name):
        # Remove population with given code name
        del self.pops[pop_name]

        for interaction in self.transfers+self.interpops:
            interaction.pops.remove(pop_name)
            for k in list(interaction.ts.keys()):
                if k[0] == pop_name or k[1] == pop_name:
                    del interaction.ts[k]

        for tdve in self.tdve.values():
            for k in list(tdve.ts.keys()):
                if k == pop_name:
                    del tdve.ts[k]

    def add_transfer(self,code_name,full_name):
        for transfer in self.transfers:
            assert code_name != transfer.code_name, 'Transfer with name "%s" already exists' % (code_name)
        self.transfers.append(TimeDependentConnections(code_name, full_name, self.tvec, list(self.pops.keys()), type='transfer', ts=None))

    def remove_transfer(self,code_name):
        names = [x.code_name for x in self.transfers]
        idx = names.index(code_name)
        del self.transfers[idx]

    # NB. Differences in the model will only happen if the model knows what to do with the new interaction
    def add_interaction(self,code_name,full_name):
        for interaction in self.interpops:
            assert code_name != interaction.code_name, 'Interaction with name "%s" already exists' % (code_name)
        interpop = TimeDependentConnections(code_name, full_name, self.tvec, list(self.pops.keys()), type='interaction', ts=None)
        self.interpops.append(interpop)
        return interpop

    def remove_interaction(self,code_name):
        names = [x.code_name for x in self.interpops]
        idx = names.index(code_name)
        del self.interpops[idx]

    def _read_pops(self, sheet):
        # TODO - can modify _read_pops() and _write_pops() if there are more population attributes
        tables = read_tables(sheet)
        assert len(tables) == 1, 'Population Definitions page should only contain one table'

        self.pops = sc.odict()
        assert tables[0][0][0].value.strip().lower() == 'abbreviation'
        assert tables[0][0][1].value.strip().lower() == 'full name'

        for row in tables[0][1:]:
            if row[0].value.strip().lower() == 'all':
                raise AtomicaException('A population was named "all", which is a reserved keyword and cannot be used as a population name')
            self.pops[row[0].value.strip()] = {'label':row[1].value.strip()}

    def _write_pops(self):
        # Writes the 'Population Definitions' sheet
        sheet = self._book.add_worksheet("Population Definitions")
        sheet.set_tab_color('#FFC000') # this tab is orange
        widths = dict()

        current_row = 0
        sheet.write(current_row, 0, 'Abbreviation', self._formats["center_bold"])
        update_widths(widths,0,'Abbreviation')
        sheet.write(current_row, 1, 'Full Name', self._formats["center_bold"])
        update_widths(widths,1,'Abbreviation')

        for name,content in self.pops.items():
            current_row += 1
            sheet.write(current_row, 0, name)
            update_widths(widths, 0, name)
            sheet.write(current_row, 1, content['label'])
            update_widths(widths, 1, content['label'])
            self._references[name] = "='%s'!%s" % (sheet.name,xlrc(current_row,0,True,True))
            self._references[content['label']] = "='%s'!%s" % (sheet.name,xlrc(current_row,1,True,True)) # Reference to the full name

        apply_widths(sheet,widths)

    def _read_transfers(self,sheet):
        tables = read_tables(sheet)
        assert len(tables)%3==0, 'There should be 3 subtables for every transfer'
        self.transfers = []
        for i in range(0,len(tables),3):
            self.transfers.append(TimeDependentConnections.from_tables(tables[i:i+3],'transfer'))
        return

    def _write_transfers(self):
        # Writes a sheet for every transfer

        # Skip if no transfers
        if not self.transfers:
            return

        sheet = self._book.add_worksheet("Transfers")
        sheet.set_tab_color('#808080')
        sheet.hide()
        widths = dict()
        next_row = 0
        for transfer in self.transfers:
            next_row = transfer.write(sheet,next_row,self._formats,self._references,widths)
        apply_widths(sheet,widths)

    def _read_interpops(self,sheet):
        tables = read_tables(sheet)
        assert len(tables)%3==0, 'There should be 3 subtables for every transfer'
        self.interpops = []
        for i in range(0,len(tables),3):
            self.interpops.append(TimeDependentConnections.from_tables(tables[i:i+3],'interaction'))
        return

    def _write_interpops(self):
        # Writes a sheet for every interaction

        # Skip if no interpops
        if not self.interpops:
            return
        sheet = self._book.add_worksheet("Interactions")
        sheet.set_tab_color('#808080')
        widths = dict()
        next_row = 0
        for interpop in self.interpops:
            next_row = interpop.write(sheet,next_row,self._formats,self._references,widths)
        apply_widths(sheet,widths)

    def _write_tdve(self):
        # Writes several sheets, one for each custom page specified in the Framework
        widths = dict()

        for sheet_name,code_names in self.tdve_pages.items():
            sheet = self._book.add_worksheet(sheet_name)
            next_row = 0
            has_editable_content = False
            for code_name in code_names:
                has_editable_content = has_editable_content or (not self.tdve[code_name].has_data) # there is editable content if any TDVE is missing data, so blue cells are present
                next_row = self.tdve[code_name].write(sheet,next_row,self._formats,self._references,widths)

            if has_editable_content:
                sheet.set_tab_color('#92D050')
            else:
                sheet.set_tab_color('#808080')

        for sheet_name in self.tdve_pages.keys():
            apply_widths(self._book.get_worksheet_by_name(sheet_name),widths)



# Program data maps to a program databook
# On construction, we first make some blank data, and then we write a databook in the same way as if we actually had
# data values
            
## QUESTION, SHOULD THIS JUST BE A PROGSET?
## IE, PROGSETS THEMSELVES HAVE TO AND FROM SPREADSHEET METHODS??
## TRY TO AVOID TOO MUCH DATA CONVERSION            
class ProgramData(object):
    def __init__(self, pops=None, comps=None, progs=None, pars=None, data=None, data_start=None, data_end=None, blh_effects=False):

#        self.book = None
#        self.formats = None
#
        self.pops = pops
        self.comps = comps
        self.progs = progs if isinstance(progs,dict) else sc.odict()
        self.pars = pars.values() if isinstance(pars,dict) else None
        self.blh_effects = blh_effects

        self.data_start = data_start if data_start is not None else 2015.0 # WARNING, remove
        self.data_end = data_end if data_end is not None else 2018.0 # WARNING, remove
        self.prog_range = None
        self.data_range = range(int(self.data_start), int(self.data_end + 1))
        self.data = data

        # Internal storage used with methods while writing
        self._formats = None
        self._book = None
        self._references = None

    def __repr__(self):
        output = sc.desc(self)
        return output


    def to_spreadsheet(self):
        # Return a AtomicaSpreadsheet with the contents of this Workbook
        f = io.BytesIO() # Write to this binary stream in memory

        self._book = xw.Workbook(f)
        self._formats = standard_formats(self._book)
        self._references = {} # Reset the references dict

        self._write_targeting(self.data)
        self._write_costcovdata(self.data)
        self._write_covoutdata(self.data)

        self._book.close()

        # Dump the file content into a ScirisSpreadsheet
        spreadsheet = AtomicaSpreadsheet(f)

        # Return the spreadsheet
        return spreadsheet


    @staticmethod
    def from_spreadsheet(spreadsheet):
        # Read in program data from a spreadsheet
        self = ProgramData()
        data=sc.odict()

        if isinstance(spreadsheet,string_types):
            spreadsheet = AtomicaSpreadsheet(spreadsheet)

        workbook = xlrd.open_workbook(file_contents=spreadsheet.get_file().read()) # Open workbook

        sheetdata = workbook.sheet_by_name('Program targeting') # Load 
        colindices = []

        for row in range(sheetdata.nrows):     
            # Get data
            thesedata = sheetdata.row_values(row, start_colx=2) 
    
            # Get metadata from first row
            if row==0:
                for col in range(2,sheetdata.ncols):
                    cell_val = sheetdata.cell(row, col).value
                    if cell_val!='': colindices.append(col-1)
    
            if row==1:
                data['pops'] = thesedata[3:colindices[1]-2]
                data['comps'] = thesedata[colindices[1]-1:]
    
            else:
                if thesedata[0]:
                    progname = str(thesedata[0])
                    data['progs']['short'].append(progname)
                    data['progs']['name'].append(str(thesedata[1])) # WARNING, don't need name and short
                    data['progs']['label'].append(str(thesedata[1]))
                    data['progs']['target_pops'].append(thesedata[3:colindices[0]])
                    data['progs']['target_comps'].append(self.blank2newtype(thesedata[colindices[1]-1:],0))
                    data[progname] = sc.odict()
                    data[progname]['name'] = str(thesedata[1])
                    data[progname]['target_pops'] = thesedata[3:colindices[0]]
                    data[progname]['target_comps'] = self.blank2newtype(thesedata[colindices[1]-1:], 0)
                    data[progname]['spend'] = []
                    data[progname]['capacity'] = []
                    data[progname]['unitcost'] = sc.odict()
    
    
    def set_content(self, name=None, row_names=None, column_names=None, row_levels=None, data=None,
                    row_format='general', row_formats=None, assumption_properties=None, assumption_data=None, assumption=True):
        # Set the content
        if assumption_properties is None:
            assumption_properties = {'title': None, 'connector': 'OR', 'columns': ['Assumption']}
        self.assumption_data = assumption_data
        return sc.odict([("name",            name),
                         ("row_names",       row_names),
                         ("column_names",    column_names),
                         ("row_levels",      row_levels),
                         ("row_format",      row_format),
                         ("row_formats",     row_formats),
                         ("data",            data),
                         ("assumption_properties", assumption_properties),
                         ("assumption_data", assumption_data),
                         ("assumption",      assumption)])
        

    def getyears(sheetdata):
        ''' Get years from a worksheet'''
        years = [] # Initialize data years
        for col in range(sheetdata.ncols):
            thiscell = sheetdata.cell_value(1,col) # 1 is the 2nd row which is where the year data should be
            if thiscell=='' and len(years)>0: #  We've gotten to the end
                lastdatacol = col # Store this column number
                break # Quit
            elif thiscell != '': # Nope, more years, keep going
                years.append(float(thiscell)) # Add this year
        
        return lastdatacol, years
       
       
    def blank2newtype(thesedata, newtype=None):
        ''' Convert a blank entry to another type, e.g. nan, None or zero'''
        if newtype is None or newtype=='nan': newval = np.nan # For backward compatability
        elif newtype=='None': newval = None
        elif newtype=='zero': newval = 0
        elif sc.isnumber(newtype): newval = newtype
        else: 
            errormsg = 'Cannot convert blanks to type %s, can only convert to types [''nan'', ''None'', ''zero''] or numbers' % (type(newtype)) 
            raise AtomicaException(errormsg)
        return [newval if thisdatum=='' else thisdatum for thisdatum in thesedata ]
        
    
    def validatedata(thesedata, sheetname, thispar, row, checkupper=False, checklower=True, checkblank=True, startcol=0):
        ''' Do basic validation on the data: at least one point entered, between 0 and 1 or just above 0 if checkupper=False '''
        
        result = sc.odict()
        result['isvalid'] = 1
        # Check that only numeric data have been entered
        for column,datum in enumerate(thesedata):
            if not sc.isnumber(datum):
                errormsg = 'Invalid entry in sheet "%s", parameter "%s":\n' % (sheetname, thispar) 
                errormsg += 'row=%i, column=%s, value="%s"\n' % (row+1, xlrd.colname(column+startcol), datum)
                errormsg += 'Be sure all entries are numeric'
                if ' ' or '\t' in datum: errormsg +=' (there seems to be a space or tab)'
                raise AtomicaException(errormsg)
        
        # Now check integrity of data itself
        validdata = np.array(thesedata)[~np.isnan(thesedata)]
        if len(validdata):
            valid = np.array([True]*len(validdata)) # By default, set everything to valid
            if checklower: valid *= np.array(validdata)>=0
            if checkupper: valid *= np.array(validdata)<=1
            if not valid.all():
                invalid = validdata[valid==False]
                errormsg = 'Invalid entry in sheet "%s", parameter "%s":\n' % (sheetname, thispar) 
                errormsg += 'row=%i, invalid="%s", values="%s"\n' % (row+1, invalid, validdata)
                errormsg += 'Be sure that all values are >=0 (and <=1 if a probability)'
                result['isvalid'] = 0
                result['errormsg'] = errormsg
        elif checkblank: # No data entered
            errormsg = 'No data or assumption entered for sheet "%s", parameter "%s", row=%i' % (sheetname, thispar, row) 
            result['isvalid'] = 0
            result['errormsg'] = errormsg
        return result

    
    def _write_targeting(self, data):
        # Generate targeting sheet
        sheet = self._book.add_worksheet('Program targeting')

        for item in items:
            if type(item) is dict:
                item_name = item['name']
                short_name = item.get('short', item_name)
                male = item.get('male', False)
                female = item.get('female', False)
                age_from = item.get('age_from',15)
                age_to = item.get('age_to',49)
        # Set column width
        sheet.set_column(2, 2, 15)
        sheet.set_column(3, 3, 40)
        sheet.set_column(6, 6, 12)
        sheet.set_column(7, 7, 16)
        sheet.set_column(8, 8, 16)
        sheet.set_column(9, 9, 12)
        current_row = 0

        # Write descriptions of targeting
        sheet.write(0, 5, "Targeted to (populations)", self._formats["bold"])
        sheet.write(0, 6 + len(self.pops), "Targeted to (compartments)", self._formats["bold"])

        # Write populations and compartments for targeting
        coded_params = []
        for item in self.progs:
            if type(item) is dict:
                name = item['name']
                short = item['short']
                target_pops = [''] + ['' for popname in self.pops]
                target_comps = [''] + ['' for comp in self.comps]
            coded_params.append([short, name] + target_pops + target_comps)

        # Make column names
        column_names = ['Short name', 'Long name', ''] + self.pops + [''] + self.comps
        content = self.set_content(row_names=range(1,len(self.progs)+1),
                                   column_names=column_names,
                                   data=coded_params,
                                   assumption=False)
        
        self.prog_range = ProgramEntry(sheet=sheet, first_row=current_row, content=content)
        current_row = self.prog_range.emit(self._formats, rc_title_align='left')
        self.ref_prog_range = self.prog_range.param_refs()


    def _write_costcovdata(self):
        # Generate cost-coverage sheet
        sheet = self._book.add_worksheet('Spending data')

        current_row = 0
        sheet.set_column('C:C', 20)
        row_levels = ['Total spend', 'Capacity constraints', 'Unit cost: best', 'Unit cost: low',
                      'Unit cost: high']
        content = self.set_content(row_names=self.ref_prog_range,
                                   column_names=range(int(self.data_start), int(self.data_end + 1)),
                                   row_formats=['general']*5,
                                   row_levels=row_levels)

        the_range = ProgramEntry(sheet=sheet, first_row=current_row, content=content)
        current_row = the_range.emit(self._formats)

    def _write_covoutdata(self):
        # Generate coverage-outcome sheet
        sheet = self._book.add_worksheet('Program effects')

        current_row = 0
        sheet.set_column(1, 1, 30)
        sheet.set_column(2, 2, 12)
        sheet.set_column(3, 3, 12)
        sheet.set_column(4, 4, 12)
        sheet.set_column(5, 5, 12)
        sheet.set_column(6, 6, 2)

        row_levels = []
        for p in self.pops:
            if self.blh_effects:
                row_levels.extend([p + ': best', p + ': low', p + ': high'])
            else: row_levels.extend([p])

        assumption_properties = {'title': 'Value for a person covered by this program alone:',
                                 'connector': '',
                                 'columns': self.ref_prog_range}

        content = self.set_content(row_names=self.pars,
                                   column_names=['Value if none of the programs listed here are targeting this parameter', 'Coverage interation', 'Impact interaction'],
                                   row_format='general',
                                   assumption_properties=assumption_properties,
                                   row_levels=row_levels)

        the_range = ProgramEntry(sheet=sheet, first_row=current_row, content=content)
        current_row = the_range.emit(self._formats, rc_title_align='left')




def make_progbook(filename, pops, comps, progs, pars, data_start=None, data_end=None, blh_effects=False):
    """ Generate the Atomica programs spreadsheet """

    # An integer argument is given: just create a pops dict using empty entries
    if sc.isnumber(pops):
        npops = pops
        pops = []  # Create real pops list
        for p in range(npops):
            pops.append('Pop %i' % (p + 1))

    if sc.isnumber(comps):
        ncomps = comps
        comps = []  # Create real compartments list
        for p in range(ncomps):
            pops.append('Comp %i' % (p + 1))

    if sc.isnumber(progs):
        nprogs = progs
        progs = []  # Create real pops list
        for p in range(nprogs):
            progs.append({'short': 'Prog %i' % (p + 1), 'name': 'Program %i' % (p + 1)})

    book = ProgramData(pops=pops, comps=comps, progs=progs, pars=pars, data_start=data_start, data_end=data_end, blh_effects=blh_effects)
    ss = book.to_spreadsheet()
    ss.save(filename)
    return filename
