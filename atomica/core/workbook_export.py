from .system import SystemSettings as SS
from .structure_settings import FrameworkSettings as FS
from .structure_settings import DataSettings as DS
from .excel import ExcelSettings as ES

from .system import logger, AtomicaException, accepts, display_name
from .excel import create_standard_excel_formats, create_default_format_variables, create_value_entry_block
from .structure_settings import DetailColumns, TableTemplate, ConnectionMatrix, TimeDependentValuesEntry, \
    IDType, IDRefType, SwitchType, QuantityFormatType, TimeDependentConnections, write_matrix
from .workbook_utils import WorkbookTypeException, get_workbook_page_keys, get_workbook_page_spec, \
    get_workbook_item_type_specs, get_workbook_item_specs
from .structure import get_quantity_type_list, TimeSeries

import sciris.core as sc
import xlsxwriter as xw
from xlsxwriter.utility import xl_rowcol_to_cell as xlrc
import numpy as np
from copy import copy
from .excel import AtomicaSpreadsheet, standard_formats
import collections
import io
import xlrd
import openpyxl

class KeyUniquenessException(AtomicaException):
    def __init__(self, key, dict_type, **kwargs):
        if key is None:
            message = ("Key uniqueness failure. A key is used more than once in '{0}'.".format(dict_type))
        else:
            message = ("Key uniqueness failure. Key '{0}' is used more than once in '{1}'.".format(key, dict_type))
        super(KeyUniquenessException, self).__init__(message, **kwargs)


class InvalidReferenceException(AtomicaException):
    def __init__(self, item_type, attribute, ref_item_type, ref_attribute, **kwargs):
        message = ("Workbook construction failed when item '{0}', attribute '{1}', attempted to reference nonexistent "
                   "values, specifically item '{2}', attribute '{3}'. It is possible the referenced attribute values "
                   "are erroneously scheduled to be created later.".format(item_type, attribute,
                                                                           ref_item_type, ref_attribute))
        super(InvalidReferenceException, self).__init__(message, **kwargs)


class WorkbookInstructions(object):
    """ An object that stores instructions for how many items should be created during workbook construction. """

    def __init__(self, workbook_type=None):
        """ Initialize instructions that detail how to construct a workbook. """
        # Every relevant item must be included in a dictionary that lists how many should be created.
        self.num_items = sc.odict()
        if workbook_type == SS.STRUCTURE_KEY_FRAMEWORK:
            item_type_specs = FS.ITEM_TYPE_SPECS
        elif workbook_type == SS.STRUCTURE_KEY_DATA:
            item_type_specs = DS.ITEM_TYPE_SPECS
        else:
            raise WorkbookTypeException(workbook_type)
        for item_type in item_type_specs:
            if item_type_specs[item_type]["instruction_allowed"]:
                self.num_items[item_type] = item_type_specs[item_type]["default_amount"]
        self.data_start = 2000.0
        self.data_end = 2020.0
        self.data_dt = 1.0

    @property
    def tvec(self):
        """ A time vector associated with workbook instructions, typically only used when generating databooks. """
        return np.arange(self.data_start, self.data_end + self.data_dt / 2, self.data_dt)

    @accepts(str, int)
    def update_number_of_items(self, item_type, number):
        """ Overwrite the number of items that will be constructed for the template workbook. """
        try:
            self.num_items[item_type] = number
        except KeyError:
            raise AtomicaException("An attempted update of workbook instructions to produce '{0}' "
                                   "instances of item type '{1}' failed.".format(number, item_type))

    def update_time_vector(self, data_start=None, data_end=None, data_dt=None):
        """ Update the attributes that determine a time vector for workbook instructions. """
        if data_start is not None:
            self.data_start = data_start
        if data_end is not None:
            self.data_end = data_end
        if data_dt is not None:
            self.data_dt = data_dt


def make_instructions(framework=None, data=None, instructions=None, workbook_type=None):
    """
    Generates instructions that detail the number of items pertinent to workbook construction processes.
    If instructions are already provided, they will be used regardless of the presence of any structure.
    Otherwise, if a ProjectFramework or ProjectData structure is available, this will be used in filling out a workbook.
    If relevant structure is unavailable, a default set of instructions will be produced instead.
    This function returns a boolean tag indicating whether to use instructions or not according to the above logic.
    """
    use_instructions = False
    # Check if framework/data is missing for the relevant workbook. If so, use instructions.
    if workbook_type == SS.STRUCTURE_KEY_FRAMEWORK:
        if framework is None:
            use_instructions = True
    elif workbook_type == SS.STRUCTURE_KEY_DATA:
        if data is None:
            use_instructions = True
    else:
        raise WorkbookTypeException(workbook_type)
    # If no instructions are provided, generate a default, even if they will not be used.
    if instructions is None:
        instructions = WorkbookInstructions(workbook_type=workbook_type)
    else:
        use_instructions = True  # If instructions are provided, they must be used.
    return instructions, use_instructions


def create_attribute_cell_content(worksheet, row, col, attribute, item_type, item_type_specs, item_number,
                                  alt_content=None, alt_condition=None,
                                  formats=None, format_key=None, temp_storage=None):
    """ Write default content into the cell of a worksheet corresponding to an attribute of an item. """

    # Determine attribute information and prepare for content production.
    attribute_spec = item_type_specs[item_type]["attributes"][attribute]
    if temp_storage is None:
        temp_storage = sc.odict()
    if formats is None:
        raise AtomicaException("Excel formats have not been passed to workbook table construction.")
    if format_key is None:
        format_key = ES.FORMAT_KEY_CENTER
    cell_format = formats[format_key]

    # Set up default content type and reference information.
    content_type = None
    do_reference = False
    other_item_type = None
    other_attribute = None

    # Determine content type and prepare for referencing if appropriate.
    if "content_type" in attribute_spec:
        content_type = attribute_spec["content_type"]
    if isinstance(content_type, IDRefType):
        do_reference = True
        if content_type.other_item_types is not None:
            other_item_type = content_type.other_item_types[0]
            other_attribute = content_type.attribute

    # Default content is blank.
    content = ""
    sep = ""
    validation_source = None
    rc = xlrc(row, col, True, True)  # Make the reference absolute for user benefits, e.g. copying tables.

    # Content associated with standard content types is set up here.
    if isinstance(content_type, IDType):
        # Prepare for referencing if this attribute has a superitem type.
        if content_type.superitem_type is not None:
            do_reference = True
            other_item_type = content_type.superitem_type
            other_attribute = attribute
        # Name and label attributes, i.e. those of ID type, reference earlier constructions of themselves if possible.
        # This is excluded for subitems due to their complicated form of content construction.
        if not do_reference and item_type in temp_storage and attribute in temp_storage[item_type] \
                and len(temp_storage[item_type][attribute]["list_content"]) > item_number:
            do_reference = True  # Is not a reference type object but will allow one-to-one referencing to take place.
            other_item_type = item_type
            other_attribute = attribute
        # Otherwise construct content with a prefix if provided.
        else:
            content = str(item_number)  # The default content is the number of this item.
            if content_type.name_not_label:
                space = SS.DEFAULT_SPACE_NAME
                sep = SS.DEFAULT_SEPARATOR_NAME
            else:
                space = SS.DEFAULT_SPACE_LABEL
                sep = SS.DEFAULT_SEPARATOR_LABEL
            if "prefix" in attribute_spec:
                content = attribute_spec["prefix"] + space + content
    elif isinstance(content_type, SwitchType):
        validation_source = [SS.DEFAULT_SYMBOL_NO, SS.DEFAULT_SYMBOL_YES]
        if content_type.default_on:
            validation_source.reverse()
        content = validation_source[0]
    elif isinstance(content_type, QuantityFormatType):
        validation_source = [""] + get_quantity_type_list(include_absolute=True, include_relative=True,
                                                          include_special=True)
        content = validation_source[0]
    content_backup = content

    # References to other content are constructed here.
    if do_reference is True:
        list_id = item_number

        # Superitem-based references link subitem attributes to corresponding superitem attributes.
        # The superitem referenced is the last one stored.
        # This is because subitem displays are meant to be created instantly after superitems.
        if isinstance(content_type, IDType) and content_type.superitem_type is not None:
            list_id = -1

        # If there is another item type to reference, proceed with referencing its ID.
        if other_item_type is not None:
            try:
                stored_refs = temp_storage[other_item_type][other_attribute]
            except KeyError:
                raise InvalidReferenceException(item_type=item_type, attribute=attribute, ref_item_type=other_item_type,
                                                ref_attribute=other_attribute)

            # For one-to-one referencing, do not create table content that goes beyond length of the referenced table.
            if len(stored_refs["list_content"]) > list_id:
                content_page = ""
                if not stored_refs["page_title"] == worksheet.name:
                    content_page = "'{0}'!".format(stored_refs["page_title"])
                ref_content = "={0}{1}".format(content_page, stored_refs["list_cell"][list_id])
                ref_content_backup = stored_refs["list_content_backup"][list_id]

                if isinstance(content_type, IDType) and content_type.superitem_type is not None:
                    content = "=CONCATENATE({0},\"{1}\")".format(ref_content.lstrip("="), sep + content)
                    content_backup = ref_content_backup + sep + content_backup
                else:
                    content = ref_content
                    content_backup = ref_content_backup

        # If the content is marked to reference its own item type, append the ID to current content.
        # This reference should be to an ID of the same item a row ago.
        if isinstance(content_type, IDRefType) and content_type.self_referencing and item_number > 0:
            list_id = item_number - 1
            try:
                stored_refs = temp_storage[item_type][content_type.attribute]
            except KeyError:
                raise InvalidReferenceException(item_type=item_type, attribute=attribute, ref_item_type=item_type,
                                                ref_attribute=content_type.attribute)
            content_page = ""
            if not stored_refs["page_title"] == worksheet.name:
                content_page = "'{0}'!".format(stored_refs["page_title"])
            ref_content = "={0}{1}".format(content_page, stored_refs["list_cell"][list_id])
            ref_content_backup = stored_refs["list_content_backup"][list_id]
            if content == "":
                content = ref_content
                content_backup = ref_content_backup
            else:
                if content.startswith("="):
                    content = content.lstrip("=")
                else:
                    content = "\"" + content + "\""
                content = "=CONCATENATE({0},\"{1}\",{2})".format(content, ES.LIST_SEPARATOR, ref_content.lstrip("="))
                content_backup = content_backup + ES.LIST_SEPARATOR + ref_content_backup

    # If content is of ID type, store it for referencing by other relevant attributes later.
    if isinstance(content_type, IDType):
        if item_type not in temp_storage:
            temp_storage[item_type] = {}
        if attribute not in temp_storage[item_type]:
            temp_storage[item_type][attribute] = {"list_content": [], "list_content_backup": [], "list_cell": []}
        # Make sure the attribute does not already have stored values associated with it.
        if not len(temp_storage[item_type][attribute]["list_content"]) > item_number:
            temp_storage[item_type][attribute]["list_content"].append(content)
            temp_storage[item_type][attribute]["list_content_backup"].append(content_backup)
            temp_storage[item_type][attribute]["list_cell"].append(rc)
            temp_storage[item_type][attribute]["page_title"] = worksheet.name

    # Do a final check to see if the content is still blank, in which case apply a default value.
    if content == "" and content_type is not None and content_type.default_value is not None:
        default_value = content_type.default_value
        if default_value is True:
            default_value = SS.DEFAULT_SYMBOL_YES
        if default_value is False:
            default_value = SS.DEFAULT_SYMBOL_NO
        content = default_value

    # Modify content for optional conditions.
    if alt_condition is not None:
        content = "=IF({0},{1},{2})".format(alt_condition, alt_content, content.lstrip("="))

    # Actually write the content, using a backup value where the content is an equation and may not be calculated.
    # This lack of calculation occurs when Excel files are not opened before writing and reading phases.
    # Also validate that the cell only allows certain values.
    if isinstance(content, str) and content.startswith("="):
        worksheet.write_formula(rc, content, cell_format, content_backup)
    else:
        worksheet.write(rc, content, cell_format)
    if validation_source is not None:
        worksheet.data_validation(rc, {"validate": "list", "source": validation_source})

    return content, content_backup


def write_headers_dc(worksheet, table, start_row, start_col, item_type=None, framework=None, data=None,
                     workbook_type=None, formats=None, format_variables=None):
    if item_type is None:
        item_type = table.item_type
    item_type_specs = get_workbook_item_type_specs(framework=framework, workbook_type=workbook_type)
    item_type_spec = item_type_specs[item_type]

    if formats is None:
        raise AtomicaException("Excel formats have not been passed to workbook table construction.")
    if format_variables is None:
        format_variables = create_default_format_variables()
    orig_format_variables = sc.dcp(format_variables)
    format_variables = sc.dcp(orig_format_variables)
    revert_format_variables = False

    row, col, header_column_map = start_row, start_col, sc.odict()
    for attribute in item_type_spec["attributes"]:
        # Ignore explicitly excluded attributes or implicitly not-included attributes for table construction.
        # Item name is always in the table though.
        if not attribute == "name":
            if table.exclude_not_include == (attribute in table.attribute_list):
                continue
        attribute_spec = item_type_spec["attributes"][attribute]
        if "ref_item_type" in attribute_spec:
            _, col, sub_map = write_headers_dc(worksheet=worksheet, table=table,
                                               item_type=attribute_spec["ref_item_type"],
                                               start_row=row, start_col=col,
                                               framework=framework, data=data, workbook_type=workbook_type,
                                               formats=formats, format_variables=format_variables)
            len_map = len(header_column_map)
            len_sub_map = len(sub_map)
            header_column_map.update(sub_map)
            if not len(header_column_map) == len_map + len_sub_map:
                raise KeyUniquenessException(None, "header-column map")
        else:
            for format_variable_key in format_variables:
                if format_variable_key in attribute_spec:
                    revert_format_variables = True
                    format_variables[format_variable_key] = attribute_spec[format_variable_key]
            header = attribute_spec["header"]
            if header in header_column_map:
                raise KeyUniquenessException(header, "header-column map")
            header_column_map[header] = col
            worksheet.write(row, col, header, formats[ES.FORMAT_KEY_CENTER_BOLD])
            if "comment" in attribute_spec:
                header_comment = attribute_spec["comment"]
                worksheet.write_comment(row, col, header_comment,
                                        {"x_scale": format_variables[ES.KEY_COMMENT_XSCALE],
                                         "y_scale": format_variables[ES.KEY_COMMENT_YSCALE]})
            worksheet.set_column(col, col, format_variables[ES.KEY_COLUMN_WIDTH])
            if revert_format_variables:
                format_variables = sc.dcp(orig_format_variables)
                revert_format_variables = False
            col += 1
    row += 1
    next_row, next_col = row, col
    return next_row, next_col, header_column_map


def write_contents_dc(worksheet, table, start_row, header_column_map, item_type=None, framework=None, data=None,
                      instructions=None, workbook_type=None,
                      formats=None, temp_storage=None):
    if item_type is None:
        item_type = table.item_type
    item_type_specs = get_workbook_item_type_specs(framework=framework, workbook_type=workbook_type)
    item_type_spec = item_type_specs[item_type]
    instructions, use_instructions = make_instructions(framework=framework, data=data, instructions=instructions,
                                                       workbook_type=workbook_type)

    if temp_storage is None:
        temp_storage = sc.odict()

    row, new_row = start_row, start_row
    if use_instructions:
        for item_number in range(instructions.num_items[item_type]):
            for attribute in item_type_spec["attributes"]:
                # Ignore explicitly excluded attributes or implicitly not-included attributes for table construction.
                # Item name is always in the table though.
                if not attribute == "name":
                    if table.exclude_not_include == (attribute in table.attribute_list):
                        continue
                attribute_spec = item_type_spec["attributes"][attribute]
                if "ref_item_type" in attribute_spec:
                    sub_row = write_contents_dc(worksheet=worksheet, table=table,
                                                item_type=attribute_spec["ref_item_type"],
                                                start_row=row, header_column_map=header_column_map,
                                                framework=framework, data=data, instructions=instructions,
                                                workbook_type=workbook_type,
                                                formats=formats, temp_storage=temp_storage)
                    new_row = max(new_row, sub_row)
                else:
                    col = header_column_map[attribute_spec["header"]]
                    create_attribute_cell_content(worksheet=worksheet, row=row, col=col,
                                                  attribute=attribute, item_type=item_type,
                                                  item_type_specs=item_type_specs,
                                                  item_number=item_number, formats=formats, temp_storage=temp_storage)
            row = max(new_row, row + 1)
    next_row = row + 1  # Delimit with a blank row.
    return next_row


def write_detail_columns(worksheet, table, start_row, start_col, framework=None, data=None, instructions=None,
                         workbook_type=None, formats=None, format_variables=None, temp_storage=None):
    if temp_storage is None:
        temp_storage = sc.odict()

    row, col = start_row, start_col
    row, _, header_column_map = write_headers_dc(worksheet=worksheet, table=table, start_row=row, start_col=col,
                                                 framework=framework, data=data, workbook_type=workbook_type,
                                                 formats=formats, format_variables=format_variables)
    row = write_contents_dc(worksheet=worksheet, table=table, start_row=row, header_column_map=header_column_map,
                            framework=framework, data=data, instructions=instructions, workbook_type=workbook_type,
                            formats=formats, temp_storage=temp_storage)
    next_row, next_col = row, col
    return next_row, next_col

def write_table(worksheet, table, start_row, start_col, framework=None, data=None, instructions=None,
                workbook_type=None, formats=None, format_variables=None, temp_storage=None):
    # Check workbook type.
    if workbook_type not in [SS.STRUCTURE_KEY_FRAMEWORK, SS.STRUCTURE_KEY_DATA]:
        raise WorkbookTypeException(workbook_type)
    item_specs = get_workbook_item_specs(framework=framework, workbook_type=workbook_type)

    instructions, use_instructions = make_instructions(framework=framework, data=data, instructions=instructions,
                                                       workbook_type=workbook_type)

    if temp_storage is None:
        temp_storage = sc.odict()

    row, col = start_row, start_col
    if isinstance(table, DetailColumns):
        row, col = write_detail_columns(worksheet=worksheet, table=table, start_row=row, start_col=col,
                                        framework=framework, data=data, instructions=instructions,
                                        workbook_type=workbook_type,
                                        formats=formats, format_variables=format_variables, temp_storage=temp_storage)
    if isinstance(table, TableTemplate):
        iteration_amount = 1
        # If the connection matrix is templated...
        if table.template_item_type is not None:
            # Check if instantiation is deferred.
            # If it is, iterate for the number of template-related items specified by instructions.
            if table.template_item_key is None:
                iteration_amount = instructions.num_items[table.template_item_type]
        if isinstance(table, ConnectionMatrix):
            for iteration in range(iteration_amount):
                raise Exception('Matrix printing should go via write_matrix()')

        if isinstance(table, TimeDependentValuesEntry):
            for iteration in range(iteration_amount):
                raise Exception('TDVE printing should go via TimeDependentValuesEntry object')

    next_row, next_col = row, col
    return next_row, next_col


def write_worksheet(workbook, page_key, framework=None, data=None, instructions=None, workbook_type=None,
                    formats=None, format_variables=None, temp_storage=None):
    page_spec = get_workbook_page_spec(page_key=page_key, framework=framework, workbook_type=workbook_type)
    if len(page_spec["tables"]) == 0:
        return

    # Construct worksheet.
    page_title = page_spec["label"]
    logger.info("Creating page: {0}".format(page_title))
    worksheet = workbook.add_worksheet(page_title)

    # Propagate file-wide format variable values to page-wide format variable values.
    # Create the format variables if they were not passed in from a file-wide context.
    # Overwrite the file-wide defaults if page-based specifics are available in framework settings.
    if format_variables is None:
        format_variables = create_default_format_variables()
    else:
        format_variables = sc.dcp(format_variables)
    for format_variable_key in format_variables:
        if format_variable_key in page_spec:
            format_variables[format_variable_key] = page_spec[format_variable_key]

    # Generate standard formats if they do not exist and construct headers for the page.
    if formats is None:
        formats = create_standard_excel_formats(workbook)

    if temp_storage is None:
        temp_storage = sc.odict()

    # Iteratively construct tables.
    row, col = 0, 0
    for table in page_spec["tables"]:
        row, col = write_table(worksheet=worksheet, table=table, start_row=row, start_col=col,
                               framework=framework, data=data, instructions=instructions, workbook_type=workbook_type,
                               formats=formats, format_variables=format_variables, temp_storage=temp_storage)


# TODO: Decide on what metadata sheet should actually contain and review if data time ranges are still necessary.
def write_reference_worksheet(workbook, framework=None, data=None, instructions=None, workbook_type=None):
    """
    Creates a hidden worksheet for metadata and other values.
    These are useful to store but are not directly part of framework/data.
    """

    if workbook_type == SS.STRUCTURE_KEY_FRAMEWORK:
        return
    instructions, _ = make_instructions(framework=framework, data=data, instructions=instructions,
                                        workbook_type=workbook_type)

    # Construct worksheet.
    page_title = "metadata".title()
    logger.info("Creating page: {0}".format(page_title))
    worksheet = workbook.add_worksheet(page_title)

    # Hard code variables and values.
    worksheet.write(0, 0, "data_start")
    worksheet.write(1, 0, "data_end")
    worksheet.write(2, 0, "data_dt")
    worksheet.write(0, 1, instructions.data_start)
    worksheet.write(1, 1, instructions.data_end)
    worksheet.write(2, 1, instructions.data_dt)

    worksheet.hide()


@accepts(str)
def write_workbook(workbook_path, framework=None, data=None, instructions=None, workbook_type=None):
    page_keys = get_workbook_page_keys(framework=framework, workbook_type=workbook_type)

    logger.info("Constructing a {0}: {1}".format(display_name(workbook_type), workbook_path))

    # Construct workbook and related formats.
    sc.makefilepath(workbook_path)
    workbook = xw.Workbook(workbook_path)
    formats = create_standard_excel_formats(workbook)
    format_variables = create_default_format_variables()

    # Create a storage dictionary for values and formulae that may persist between sections.
    temp_storage = sc.odict()

    # Iteratively construct worksheets.
    for page_key in page_keys:
        write_worksheet(workbook=workbook, page_key=page_key,
                        framework=framework, data=data, instructions=instructions, workbook_type=workbook_type,
                        formats=formats, format_variables=format_variables, temp_storage=temp_storage)
    write_reference_worksheet(workbook=workbook, framework=framework, data=data,
                              instructions=instructions, workbook_type=workbook_type)
    workbook.close()

    logger.info("{0} construction complete.".format(display_name(workbook_type, as_title=True)))


# %% COMPLETELY INDEPENDENT CODE TO MAKE A SPREADSHEET FOR PROGRAMS.
# TODO: reconcile these!!!

class AtomicaFormats:
    """ the formats used in the spreadsheet """
    darkgray = '#413839'
    originalblue = '#18C1FF'
    optionalorange = '#FFA500'
    BG_COLOR = originalblue
    OPT_COLOR = optionalorange
    BORDER_COLOR = 'white'

    PERCENTAGE = 'percentage'
    RATE = 'rate'
    DECIMAL = 'decimal'
    SCIENTIFIC = 'scientific'
    NUMBER = 'number'
    GENERAL = 'general'
    OPTIONAL = 'optional'

    def __init__(self, book):
        self.formats = {}
        self.book = book
        # locked formats
        self.formats['bold'] = self.book.add_format({'bold': 1})
        self.formats['center'] = self.book.add_format({'align': 'center'})
        self.formats['center_bold'] = self.book.add_format({'bold': 1, 'align': 'center'})
        self.formats['rc_title'] = {}
        self.formats['rc_title']['right'] = {}
        self.formats['rc_title']['right']['T'] = self.book.add_format({'bold': 1, 'align': 'right', 'text_wrap': True})
        self.formats['rc_title']['right']['F'] = self.book.add_format({'bold': 1, 'align': 'right', 'text_wrap': False})
        self.formats['rc_title']['left'] = {}
        self.formats['rc_title']['left']['T'] = self.book.add_format({'bold': 1, 'align': 'left', 'text_wrap': True})
        self.formats['rc_title']['left']['F'] = self.book.add_format({'bold': 1, 'align': 'left', 'text_wrap': False})
        # unlocked formats
        self.formats['unlocked'] = self.book.add_format({'locked': 0, 'bg_color': AtomicaFormats.BG_COLOR, 'border': 1,
                                                         'border_color': AtomicaFormats.BORDER_COLOR})
        self.formats['percentage'] = self.book.add_format(
            {'locked': 0, 'num_format': 0x09, 'bg_color': AtomicaFormats.BG_COLOR, 'border': 1,
             'border_color': AtomicaFormats.BORDER_COLOR})
        self.formats['rate'] = self.book.add_format(
            {'locked': 0, 'num_format': 0x09, 'bg_color': AtomicaFormats.BG_COLOR, 'border': 1,
             'border_color': AtomicaFormats.BORDER_COLOR})
        self.formats['decimal'] = self.book.add_format(
            {'locked': 0, 'num_format': 0x0a, 'bg_color': AtomicaFormats.BG_COLOR, 'border': 1,
             'border_color': AtomicaFormats.BORDER_COLOR})
        self.formats['scientific'] = self.book.add_format(
            {'locked': 0, 'num_format': 0x0b, 'bg_color': AtomicaFormats.BG_COLOR, 'border': 1,
             'border_color': AtomicaFormats.BORDER_COLOR})
        self.formats['number'] = self.book.add_format(
            {'locked': 0, 'num_format': 0x04, 'bg_color': AtomicaFormats.BG_COLOR, 'border': 1,
             'border_color': AtomicaFormats.BORDER_COLOR})
        self.formats['general'] = self.book.add_format(
            {'locked': 0, 'num_format': 0x00, 'bg_color': AtomicaFormats.BG_COLOR, 'border': 1,
             'border_color': AtomicaFormats.BORDER_COLOR})
        self.formats['optional'] = self.book.add_format(
            {'locked': 0, 'num_format': 0x00, 'bg_color': AtomicaFormats.OPT_COLOR, 'border': 1,
             'border_color': AtomicaFormats.BORDER_COLOR})
        self.formats['info_header'] = self.book.add_format(
            {'align': 'center', 'valign': 'vcenter', 'color': '#D5AA1D', 'fg_color': '#0E0655', 'font_size': 20})
        self.formats['grey'] = self.book.add_format({'fg_color': '#EEEEEE', 'text_wrap': True})
        self.formats['orange'] = self.book.add_format({'fg_color': '#FFC65E', 'text_wrap': True})
        self.formats['info_url'] = self.book.add_format(
            {'fg_color': '#EEEEEE', 'text_wrap': True, 'color': 'blue', 'align': 'center'})
        self.formats['grey_bold'] = self.book.add_format({'fg_color': '#EEEEEE', 'bold': True})
        self.formats['merge_format'] = self.book.add_format({'bold': 1, 'align': 'center', 'text_wrap': True})

    def write_block_name(self, sheet, name, row):
        sheet.write(row, 0, name, self.formats['bold'])

    def write_rowcol_name(self, sheet, row, col, name, align='right', wrap='T'):
        sheet.write(row, col, name, self.formats['rc_title'][align][wrap])

    def write_option(self, sheet, row, col, name='OR'):
        sheet.write(row, col, name, self.formats['center_bold'])

    # special processing for bool values (to keep the content separate from representation)
    def write_unlocked(self, sheet, row, col, data, row_format='unlocked'):
        if type(data) == bool:
            bool_data = 'TRUE' if data else 'FALSE'
            sheet.write(row, col, bool_data, self.formats[row_format])
        else:
            sheet.write(row, col, data, self.formats[row_format])

    def write_empty_unlocked(self, sheet, row, col, row_format='unlocked'):
        sheet.write_blank(row, col, None, self.formats[row_format])

    def writeline(self, sheet, row, row_format='grey'):
        sheet.write_blank(row, 0, None, self.formats[row_format])
        return row + 1

    def writeblock(self, sheet, row, text, row_format='grey', row_height=None, add_line=True):
        if row_height:
            sheet.set_row(row, row_height)
        sheet.write(row, 0, text, self.formats[row_format])
        if add_line:
            return self.writeline(sheet, row + 1)
        else:
            return row + 1


class SheetRange:
    def __init__(self, first_row, first_col, num_rows, num_cols):
        self.first_row = first_row
        self.first_col = first_col

        self.num_rows = num_rows
        self.num_cols = num_cols

        self.last_row = self.first_row + self.num_rows - 1
        self.last_col = self.first_col + self.num_cols - 1

        self.start = self.get_cell_address(self.first_row, self.first_col)
        self.end = self.get_cell_address(self.last_row, self.last_col)

    def get_address(self):
        return '%s:%s' % (self.start, self.end)

    def get_cell_address(self, row, col):
        return xw.utility.xl_rowcol_to_cell(row, col, row_abs=True, col_abs=True)

    def param_refs(self, sheet_name, column_number=1):
        """ gives the list of references to the entries in the row names (which are parameters) """
        par_range = range(self.first_row, self.last_row + 1)
        return ["='%s'!%s" % (sheet_name, self.get_cell_address(row, self.first_col + column_number)) for row in
                par_range]


class TitledRange(object):
    FIRST_COL = 2
    ROW_INTERVAL = 3

    def __init__(self, sheet=None, first_row=None, content=None):
        self.sheet = sheet
        self.content = content
        first_data_col = TitledRange.FIRST_COL
        num_data_rows = len(self.content.row_names)

        if self.content.row_levels is not None:
            first_data_col += 1
            num_data_rows *= len(self.content.row_levels)
            num_data_rows += len(self.content.row_names) - 1

        self.data_range = SheetRange(first_row + 2, first_data_col, num_data_rows, len(self.content.column_names))
        self.first_row = first_row

    def num_rows(self):
        return self.data_range.num_rows + 2

    def emit(self, formats, rc_row_align='right', rc_title_align='right'):  # only important for row/col titles
        """ Emits the range and returns the new current row in the given sheet """
        # top-top headers
        formats.write_block_name(self.sheet, self.content.name, self.first_row)

        if self.content.assumption and self.first_row == 0 and self.content.assumption_properties['title'] is not None:
            formats.write_rowcol_name(sheet=self.sheet, row=self.first_row, col=self.data_range.last_col + 2,
                                      name=self.content.assumption_properties['title'], align='left', wrap='F')

        # headers
        for i, name in enumerate(self.content.column_names):
            formats.write_rowcol_name(self.sheet, self.first_row + 1, self.data_range.first_col + i, name,
                                      rc_title_align, )

        if self.content.assumption:
            for index, col_name in enumerate(self.content.assumption_properties['columns']):
                formats.write_rowcol_name(self.sheet, self.first_row + 1, self.data_range.last_col + 2 + index,
                                          col_name)

        current_row = self.data_range.first_row
        num_levels = len(self.content.row_levels) if self.content.row_levels is not None else 1

        # iterate over rows, incrementing current_row as we go
        for i, names_format in enumerate(zip(self.content.get_row_names(), self.content.get_row_formats())):
            names, row_format = names_format
            start_col = self.data_range.first_col - len(names)
            # emit row name(s)
            for n, name in enumerate(names):
                formats.write_rowcol_name(self.sheet, current_row, start_col + n, name, rc_row_align, wrap='F')
            # emit data if present
            savedata = False
            if self.content.data is not None:
                try:
                    for j, item in enumerate(self.content.data[i]):
                        formats.write_unlocked(self.sheet, current_row, self.data_range.first_col + j, item, row_format)
                    savedata = True  # It saved successfully
                except:
                    errormsg = 'WARNING, failed to save "%s" with data:\n%s' % (self.content.name, self.content.data)
                    print(errormsg)
                    savedata = False
            if not savedata:
                for j in range(self.data_range.num_cols):
                    formats.write_empty_unlocked(self.sheet, current_row, self.data_range.first_col + j, row_format)
            # emit assumption
            if self.content.assumption:
                formats.write_option(self.sheet, current_row, self.data_range.last_col + 1,
                                     name=self.content.assumption_properties['connector'])
                for index, col_name in enumerate(self.content.assumption_properties['columns']):
                    saveassumptiondata = False
                    if self.content.assumption_data is not None:
                        try:
                            assumptiondata = self.content.assumption_data[i]
                            if isinstance(assumptiondata, list):  # Check to see if it's a list
                                if len(assumptiondata) != 1:  # Check to see if it has the right length
                                    errormsg = 'WARNING, assumption "%s" appears to have the wrong length:\n%s' % (
                                        self.content.name, assumptiondata)
                                    print(errormsg)
                                    saveassumptiondata = False
                                else:  # It has length 1, it's good to go
                                    assumptiondata = assumptiondata[0]  # Just pull out the only element
                            formats.write_unlocked(self.sheet, current_row, self.data_range.last_col + 2 + index,
                                                   assumptiondata, row_format)
                            saveassumptiondata = True
                        except Exception as E:
                            errormsg = 'WARNING, failed to save assumption "%s" with data:\n%s\nError message:\n (%s)' % (
                                self.content.name, self.content.assumption_data, repr(E))
                            print(errormsg)
                            saveassumptiondata = False
                            raise E
                    if not saveassumptiondata:
                        formats.write_empty_unlocked(self.sheet, current_row, self.data_range.last_col + 2 + index,
                                                     row_format)
            current_row += 1
            if num_levels > 1 and ((i + 1) % num_levels) == 0:  # shift between the blocks
                current_row += 1
        # done! return the new current_row plus spacing
        return current_row + TitledRange.ROW_INTERVAL  # for spacing

    def param_refs(self, column_number=0):
        return self.data_range.param_refs(self.sheet.get_name(), column_number)


class AtomicaContent(object):
    """ the content of the data ranges (row names, column names, optional data and assumptions) """

    def __init__(self, name=None, row_names=None, column_names=None, row_levels=None, data=None,
                 assumption_properties=None, assumption_data=None, assumption=True):
        self.name = name
        self.row_names = row_names
        self.column_names = column_names
        self.data = data
        self.assumption = assumption
        self.row_levels = row_levels
        self.row_format = AtomicaFormats.GENERAL
        self.row_formats = None
        if assumption_properties is None:
            self.assumption_properties = {'title': None, 'connector': 'OR', 'columns': ['Assumption']}
        else:
            self.assumption_properties = assumption_properties
        self.assumption_data = assumption_data

    def get_row_names(self):
        if not self.row_levels is not None:
            return [[name] for name in self.row_names]
        else:
            return [[name, level] for name in self.row_names for level in self.row_levels]

    def get_row_formats(self):  # assume that the number of row_formats is same as the number of row_levels
        if not self.row_levels is not None:
            return [self.row_format for name in self.row_names]
        else:
            if self.row_formats is not None:
                return [row_format for name in self.row_names for row_format in self.row_formats]
            else:
                return [self.row_format for name in self.row_names for level in self.row_levels]


# %% Base class for framework file, databook and progbook.

class Workbook(object):
    def __init__(self, name):
        self.name = name
        self.sheet_names = sc.odict()

        self.book = None
        self.sheets = None
        self.formats = None
        self.current_sheet = None

    def create(self):
        # Return a AtomicaSpreadsheet with the contents of this Workbook
        f = io.BytesIO() # Write to this binary stream in memory

        self.book = xw.Workbook(f)
        self.formats = standard_formats(self.book)
        self.write_pages()    # This calls an overloadable method for nonstandard page generation.
        self.book.close()

        # Dump the file content into a AtomicaSpreadsheet
        return AtomicaSpreadsheet(f)

    def write_pages(self):
        # By default, call `self.generate_<sheetname>` for every sheet in self.sheets
        # However, overloading this function enables writing arbitrary sheets. This is important because the `generate...`
        # function does not create the sheet and some functions, like writing multiple transfer tables or user-controlled
        # TDVE pages, might create multiple sheets
        self.sheets = {}
        for name in self.sheet_names:
            self.sheets[name] = self.book.add_worksheet(self.sheet_names[name])
            self.current_sheet = self.sheets[name]
            getattr(self, "generate_{0}".format(name))()  # This calls the corresponding page generating function.

    def construct_other_pages(self):
        """ Method that is only typically overloaded by Databook, due to its UI-driven page organisation. """
        pass

# %% Functions shared between workbook objects.

def make_table_detail_columns(sheet, headers, content,formats=None):
    """
    Construct table for current sheet of 'file_object' that allows items and their attributes to be detailed.
    Consider each column to bey keyed with a so-called attribute key.
    Arg 'headers' must be an odict mapping attribute keys with header strings.
    Arg 'content' must be a list of dicts, each mapping the same attribute keys with values per exported item.
    """
    max_string_lengths = dict()  # Store maximum string length per column in case other tables need it for autofit.
    extra_width = 1  # Stretch autofitted columns by this extra amount.
    for attribute_id, attribute in enumerate(headers):
        header_row = 0
        col = attribute_id
        # Note char length of headers for autofit purposes.
        max_string_lengths[col] = len(str(headers[attribute]))
        # Write headers.
        sheet.write(header_row, col, headers[attribute], formats["center_bold"])
        for item_id, item in enumerate(content):
            # If item details are not specified for a column attribute, move to next item.
            if attribute not in item:
                continue
            else:
                val = item[attribute]
            # Convert unusual content types to strings.
            if isinstance(val, list):
                val = ", ".join(val)
            elif val is True:
                val = "y"
            elif val is False:
                val = "n"
            # Update maximum char length of column if necessary.
            max_string_lengths[col] = max(len(str(val)), max_string_lengths[col])
            # Write content.
            row = item_id + 1
            sheet.write(row, col, val, formats["center"])
        # Approximate column autofit.
        sheet.set_column(col, col, max_string_lengths[col] + extra_width)

    # Return next row for subsequent tables, as well a dictionary of maximum string lengths by column.
    next_row = len(content) + 2
    return next_row, max_string_lengths

# %% Framework file exports.

class FrameworkFile(Workbook):
    def __init__(self, name, datapages, comps, characs, interpops, pars):
        super(FrameworkFile, self).__init__(name='')
        self.datapages = datapages
        self.comps = comps
        self.characs = characs
        self.interpops = interpops
        self.pars = pars
        self.references = {}

    def write_pages(self):
        self.write_datapage()
        self.write_comp()
        self.write_link()
        self.write_charac()
        self.write_interpop()
        self.write_par()

    def write_datapage(self):
        sheet = self.book.add_worksheet("Custom Databook Pages")

        sheet_headers = sc.odict([
            ("name", "Datasheet Code Name"),
            ("label", "Datasheet Title")
        ])
        make_table_detail_columns(sheet=sheet, headers=sheet_headers, content=self.datapages,formats=self.formats)

    def write_comp(self):
        sheet = self.book.add_worksheet("Compartments")

        sheet_headers = sc.odict([
            ("name", "Code Name"),
            ("label", "Display Name"),
            ("is_source", "Is Source"),
            ("is_sink", "Is Sink"),
            ("is_junction", "Is Junction"),
            ("default_value", "Default Value"),
            ("setup_weight", "Setup Weight"),
            ("can_calibrate", "Can Calibrate"),
            ("datapage", "Databook Page"),
            ("datapage_order", "Databook Order"),
            ("cascade_stage", "Cascade Stage")
        ])
        make_table_detail_columns(sheet=sheet, headers=sheet_headers, content=self.comps,formats=self.formats)

        for i,comp in enumerate(self.comps):
            self.references[comp['name']] = "='%s'!%s" % (sheet.name,xlrc(i+1,0,True,True))

    def write_link(self):
        sheet = self.book.add_worksheet("Transitions")

        nodes = [x['name'] for x in self.comps]
        entries = {}
        for par in self.pars:
            if 'links' in par:
                for link in par['links']:
                    entries[link] = par['name']

        write_matrix(sheet, 0, nodes, entries, self.formats, references=self.references)

    def write_charac(self):
        sheet = self.book.add_worksheet("Characteristics")

        sheet_headers = sc.odict([
            ("name", "Code Name"),
            ("label", "Display Name"),
            ("includes", "Components"),
            ("denominator", "Denominator"),
            ("default_value", "Default Value"),
            ("setup_weight", "Setup Weight"),
            ("can_calibrate", "Can Calibrate"),
            ("datapage", "Databook Page"),
            ("datapage_order", "Databook Order"),
            ("cascade_stage", "Cascade Stage")
        ])
        make_table_detail_columns(sheet=sheet, headers=sheet_headers, content=self.characs,formats=self.formats)

    def write_interpop(self):
        sheet = self.book.add_worksheet("Interactions")

        sheet_headers = sc.odict([
            ("name", "Code Name"),
            ("label", "Display Name"),
            ("default_value", "Default Value")
        ])
        make_table_detail_columns(sheet=sheet, headers=sheet_headers, content=self.interpops,formats=self.formats)

    def write_par(self):
        sheet = self.book.add_worksheet("Parameters")

        sheet_headers = sc.odict([
            ("name", "Code Name"),
            ("label", "Display Name"),
            ("format", "Format"),
            ("default_value", "Default Value"),
            ("min", "Minimum Value"),
            ("max", "Maximum Value"),
            ("func", "Function"),
            ("is_impact", "Is Impact"),
            ("can_calibrate", "Can Calibrate"),
            ("datapage", "Databook Page"),
            ("datapage_order", "Databook Order")
        ])
        make_table_detail_columns(sheet=sheet, headers=sheet_headers, content=self.pars,formats=self.formats)


def make_framework_file(filename, datapages, comps, characs, interpops, pars, framework=None):
    """ Generate the Atomica framework file as an Excel workbook. """

    item_types = ["datapage", "comp", "charac", "interpop", "par"]
    item_type_inputs = [datapages, comps, characs, interpops, pars]
    # Iterate through item type to set up lists of items and their details stored elementwise in dicts.
    # This will form the content of the Excel workbook.
    for j in range(len(item_types)):
        item_type = item_types[j]
        item_type_input = item_type_inputs[j]

        item_details = []
        # If framework is passed in, each element of 'item_details' is just their dictionary of specifications.
        # Item names must be inserted into each dict though, as they are originally treated as keys for the dict.
        if framework is not None:
            item_specs = copy(framework.specs[item_type])
            # TODO: Shallow copy is safe for specs dictionaries that are not nested.
            #       This may change, depending on population attribute/options and other potential future item types.
            for item_key in item_specs:
                item_specs[item_key].update({"name": item_key})
                item_details.append(item_specs[item_key])

        # If an integer argument is given, just create (or extend) a list using empty entries.
        # TODO: Consider hard-coding item attributes somewhere common where this code and header-names can access them.
        if sc.isnumber(item_type_input):
            num_items = item_type_input
            # Keep if clause out of 'for' loops.
            if item_type == "datapage":
                for k in range(num_items):
                    item_details.append({"name": "sh_{0}".format(k), "label": "Custom Databook Sheet {0}".format(k)})
            elif item_type == "comp":
                for k in range(num_items):
                    item_details.append({"name": "comp_{0}".format(k), "label": "Compartment {0}".format(k),
                                         "is_source": False, "is_sink": False, "is_junction": False,
                                         "default_value": None, "setup_weight": 1, "can_calibrate": True,
                                         "datapage": None, "datapage_order": None, "cascade_stage": None})
            elif item_type == "charac":
                for k in range(num_items):
                    # TODO: Maybe consider cross-referencing 'includes' to comps/characs to exemplify default.
                    item_details.append({"name": "charac_{0}".format(k), "label": "Characteristic {0}".format(k),
                                         "includes": list(), "denominator": None,
                                         "default_value": None, "setup_weight": 1, "can_calibrate": True,
                                         "datapage": None, "datapage_order": None, "cascade_stage": None})
            elif item_type == "interpop":
                for k in range(num_items):
                    item_details.append({"name": "interpop_{0}".format(k), "label": "Interaction {0}".format(k),
                                         "default_value": 1})
            elif item_type == "par":
                for k in range(num_items):
                    item_details.append({"name": "par_{0}".format(k), "label": "Parameter {0}".format(k),
                                         "format": None, "default_value": None, "min": None, "max": None,
                                         "can_calibrate": True, "datapage": None, "datapage_order": None})

        # TODO: Ensure that non-integer function inputs are of the right type when using them as item details.
        #       Alternatively, just disable the else case if manual spec construction is redundant.
        else:
            item_details.extend(item_type_input)
        item_type_inputs[j] = item_details

    # Construct the actual workbook from generated content.
    book = FrameworkFile(filename, *item_type_inputs)
    ss = book.create()  # This is a AtomicaSpreadsheet that can be stored in the FE database
    ss.save(filename)

    return filename


# %% Databook exports.



# %% Program spreadsheet exports.

class ProgramSpreadsheet(Workbook):
    def __init__(self, filename, pops, comps, progs, pars, data_start=None, data_end=None, blh_effects=False):
        super(ProgramSpreadsheet, self).__init__(name=filename)
        self.sheet_names = sc.odict([
            ('targeting', 'Program targeting'),
            ('costcovdata', 'Spending data'),
            ('covoutdata', 'Program effects'),
        ])

        self.pops = pops
        self.comps = comps
        self.progs = progs
        self.pars = pars
        self.blh_effects = blh_effects

        self.data_start = data_start if data_start is not None else 2015.0 # WARNING, remove
        self.data_end = data_end if data_end is not None else 2018.0 # WARNING, remove
        self.prog_range = None
        self.ref_pop_range = None
        self.data_range = range(int(self.data_start), int(self.data_end + 1))

        self.npops = len(pops)
        self.nprogs = len(progs)

    def to_spreadsheet(self):
        # Return a AtomicaSpreadsheet with the contents of this Workbook
        f = io.BytesIO() # Write to this binary stream in memory

        self.book = xw.Workbook(f)
        self.formats = AtomicaFormats(self.book)
        self.write_pages()    # This calls an overloadable method for nonstandard page generation.
        self.book.close()

        # Dump the file content into a AtomicaSpreadsheet
        return AtomicaSpreadsheet(f)

    def generate_targeting(self):
        self.current_sheet.set_column(2, 2, 15)
        self.current_sheet.set_column(3, 3, 40)
        self.current_sheet.set_column(6, 6, 12)
        self.current_sheet.set_column(7, 7, 16)
        self.current_sheet.set_column(8, 8, 16)
        self.current_sheet.set_column(9, 9, 12)
        current_row = 0

        coded_params = []
        for item in self.progs:
            if type(item) is dict:
                name = item['name']
                short = item['short']
                target_pops = [''] + ['' for popname in self.pops]
                target_comps = [''] + ['' for comp in self.comps]
            coded_params.append([short, name] + target_pops + target_comps)

        # Hard-coded writing of target descriptions in sheet.
        self.current_sheet.write(0, 5, "Targeted to (populations)", self.formats.formats["center_bold"])
        self.current_sheet.write(0, 6 + len(self.pops), "Targeted to (compartments)",
                                 self.formats.formats["center_bold"])

        column_names = ['Short name', 'Long name', ''] + self.pops + [''] + self.comps
        content = AtomicaContent(name='',
                                 row_names=range(1, len(self.progs) + 1),
                                 column_names=column_names,
                                 data=coded_params,
                                 assumption=False)
        self.prog_range = TitledRange(sheet=self.current_sheet, first_row=current_row, content=content)
        current_row = self.prog_range.emit(self.formats, rc_title_align='left')
        self.ref_prog_range = self.prog_range

    def generate_costcovdata(self):
        current_row = 0
        self.current_sheet.set_column('C:C', 20)
        row_levels = ['Total spend', 'Capacity constraints', 'Unit cost: best', 'Unit cost: low',
                      'Unit cost: high']
        content = AtomicaContent(name='',
                                 row_names=self.ref_prog_range.param_refs(),
                                 column_names=range(int(self.data_start), int(self.data_end + 1)))
        content.row_formats = [AtomicaFormats.SCIENTIFIC, AtomicaFormats.GENERAL, AtomicaFormats.GENERAL,
                               AtomicaFormats.GENERAL]
        content.assumption = True
        content.row_levels = row_levels
        the_range = TitledRange(self.current_sheet, current_row, content)
        content.get_row_formats()
        current_row = the_range.emit(self.formats)

    def generate_covoutdata(self):
        current_row = 0
        self.current_sheet.set_column(1, 1, 30)
        self.current_sheet.set_column(2, 2, 12)
        self.current_sheet.set_column(3, 3, 12)
#        self.current_sheet.set_column(4, 4, 12)
#        self.current_sheet.set_column(5, 5, 2)
        self.current_sheet.set_column(4, 4, 2)

        row_levels = []
        for p in self.pops:
            if self.blh_effects:
                row_levels.extend([p + ': best', p + ': low', p + ': high'])
            else: row_levels.extend([p])
        content = AtomicaContent(row_names=self.pars,
                                 column_names=['Value if none of the programs listed here are targeting this parameter'])
        content.row_format = AtomicaFormats.GENERAL
        content.row_levels = row_levels

        assumption_properties = {'title': 'Value for a person covered by this program alone:',
                                 'connector': '',
                                 'columns': self.ref_prog_range.param_refs()}

        content.assumption_properties = assumption_properties
        the_range = TitledRange(self.current_sheet, current_row, content)
        current_row = the_range.emit(self.formats, rc_title_align='left')


def make_progbook(filename, pops, comps, progs, pars, data_start=None, data_end=None, blh_effects=False):
    """ Generate the Atomica programs spreadsheet """

    # An integer argument is given: just create a pops dict using empty entries
    if sc.isnumber(pops):
        npops = pops
        pops = []  # Create real pops list
        for p in range(npops):
            pops.append('Pop %i' % (p + 1))

    if sc.isnumber(comps):
        ncomps = comps
        comps = []  # Create real compartments list
        for p in range(ncomps):
            pops.append('Comp %i' % (p + 1))

    if sc.isnumber(progs):
        nprogs = progs
        progs = []  # Create real pops list
        for p in range(nprogs):
            progs.append({'short': 'Prog %i' % (p + 1), 'name': 'Program %i' % (p + 1)})

    book = ProgramSpreadsheet(filename=filename, pops=pops, comps=comps, progs=progs, pars=pars, data_start=data_start, data_end=data_end, blh_effects=blh_effects)
    ss = book.to_spreadsheet()
    ss.save(filename)
    return filename


# %% Tests

# from atomica.core.project import Project
# from atomica.core.framework import ProjectFramework
# from atomica.core.data import ProjectData
#
# F = ProjectFramework(filepath="frameworks/framework_tb.xlsx")
# make_framework_file("f_blug.xlsx", datapages=3, comps=5, characs=7, interpops=9, pars=11, framework=F)
# P = Project(name="test", framework=F, do_run=False)
# P.load_databook(databook_path="databooks/databook_tb.xlsx", make_default_parset=False, do_run=False)
# make_databook("d_blug.xlsx", pops=2, transfers=4, framework=P.framework, data=P.data)
