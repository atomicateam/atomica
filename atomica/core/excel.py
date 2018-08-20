# -*- coding: utf-8 -*-
"""
Atomica Excel utilities file.
Contains functionality specific to Excel input and output.
"""

from .system import AtomicaException

from xlsxwriter.utility import xl_rowcol_to_cell as xlrc
import sciris.core as sc
import io
import openpyxl
from openpyxl.comments import Comment
import numpy as np
from .structure import FrameworkSettings as FS
from six import string_types
from .system import logger

def standard_formats(workbook):
    # Add standard formatting to a workbook and return the set of format objects
    # for use when writing within the workbook

    """ the formats used in the spreadsheet """
#    darkgray = '#413839'
    originalblue = '#18C1FF'
    optionalorange = '#FFA500'
    BG_COLOR = originalblue
    OPT_COLOR = optionalorange
    BORDER_COLOR = 'white'

    formats = {}

    # Locked formats
    formats['bold'] = workbook.add_format({'bold': 1})
    formats['center'] = workbook.add_format({'align': 'center'})
    formats['center_bold'] = workbook.add_format({'bold': 1, 'align': 'center'})
    formats['rc_title'] = {}
    formats['rc_title']['right'] = {}
    formats['rc_title']['right']['T'] = workbook.add_format({'bold': 1, 'align': 'right', 'text_wrap': True})
    formats['rc_title']['right']['F'] = workbook.add_format({'bold': 1, 'align': 'right', 'text_wrap': False})
    formats['rc_title']['left'] = {}
    formats['rc_title']['left']['T'] = workbook.add_format({'bold': 1, 'align': 'left', 'text_wrap': True})
    formats['rc_title']['left']['F'] = workbook.add_format({'bold': 1, 'align': 'left', 'text_wrap': False})

    # Unlocked formats
    formats['unlocked'] = workbook.add_format({'locked': 0, 'bg_color': BG_COLOR, 'border': 1,'border_color': BORDER_COLOR})
    formats['center_unlocked'] = workbook.add_format({'align': 'center','locked': 0, 'bg_color': BG_COLOR, 'border': 1,'border_color': BORDER_COLOR})
    formats['general'] = workbook.add_format({'locked': 0, 'num_format': 0x00, 'bg_color': BG_COLOR, 'border': 1,'border_color': BORDER_COLOR})

    # Conditional formats
    formats['unlocked_boolean_true'] = workbook.add_format({'bg_color': OPT_COLOR})
    formats['unlocked_boolean_false'] = workbook.add_format({'bg_color': BG_COLOR})
    formats['not_required'] = workbook.add_format({'bg_color': '#EEEEEE','border': 1,'border_color': '#CCCCCC'})
    formats['white_bg'] = workbook.add_format({'bg_color': '#FFFFFF','border': 1,'border_color': '#CCCCCC'})
    formats['ignored'] = workbook.add_format({'pattern': 14}) # Hatched with diagonal lines - this represents a cell whose value will not be used in the model run (e.g., an assumption that also has time-specific points)
    formats['warning'] = workbook.add_format({'bg_color': '#FF0000'})
    formats['ignored_warning'] = workbook.add_format({'pattern': 14,'bg_color': '#FF0000'})

    return formats


def apply_widths(worksheet,width_dict):
    for idx,width in width_dict.items():
        worksheet.set_column(idx, idx, width*1.1 + 1)


def update_widths(width_dict,column_index,contents):
    # Keep track of the maximum length of the contents in a column
    # width_dict is a dict that is keyed by column index e.g. 0,1,2
    # and the value is the length of the longest contents seen for that column
    if width_dict is None or contents is None or not isinstance(contents,string_types):
        return

    if len(contents) == 0:
        return

    if column_index not in width_dict:
        width_dict[column_index] = len(contents)
    else:
        width_dict[column_index] = max(width_dict[column_index],len(contents))


class AtomicaSpreadsheet(object):
    ''' A class for reading and writing data in binary format, so a project contains the spreadsheet loaded '''
    # This object provides an interface for managing the contents of files (particularly spreadsheets) as Python objects
    # that can be stored in the FE database. Basic usage is as follows:
    #
    # ss = AtomicaSpreadsheet('input.xlsx') # Load a file into this object
    # f = ss.get_file() # Retrieve an in-memory file-like IO stream from the data
    # book = openpyxl.load_workbook(f) # This stream can be passed straight to openpyxl
    # book.create_sheet(...)
    # book.save(f) # The workbook can be saved back to this stream
    # ss.insert(f) # We can update the contents of the AtomicaSpreadsheet with the newly written workbook
    # ss.save('output.xlsx') # Can also write the contents back to disk
    #
    # As shown above, no disk IO needs to happen to manipulate the spreadsheets with openpyxl (or xlrd/xlsxwriter)

    def __init__(self, source=None):
        # source is a specification of where to get the data from
        # It can be anything supported by AtomicaSpreadsheet.insert() which are
        # - A filename, which will get loaded
        # - A io.BytesIO which will get dumped into this instance

        self.filename = None
        self.data = None
        self.load_date = None

        if source is not None:
            self.insert(source)

    def __repr__(self):
        output = sc.desc(self)
        return output

    def insert(self, source):
        # This function sets the `data` attribute given a file-like data source
        #
        # INPUTS:
        # - source : This contains the contents of the file. It can be
        #   - A string, which is interpreted as a filename
        #   - A file-like object like a BytesIO, the entire contents of which will be read
        #
        # This function reads a binary ile on disk and stores the content in self.data
        # It also records where the file was loaded from and the date
        if isinstance(source,io.BytesIO):
            source.flush()
            source.seek(0)
            self.data = source.read()
        else:
            filepath = sc.makefilepath(filename=source)
            self.filename = filepath
            self.load_date = sc.today()
            with open(filepath, mode='rb') as f:
                self.data = f.read()

        self.load_date = sc.today()

    def save(self, filename=None):
        # This function writes the contents of self.data to a file on disk
        if filename is None:
            if self.filename is not None:
                filename = self.filename
            else:
                raise Exception('Cannot determine filename')

        filepath = sc.makefilepath(filename=filename)
        with open(filepath, mode='wb') as f:
            f.write(self.data)
        print('Spreadsheet saved to %s.' % filepath)

        return filepath

    def get_file(self):
        # Return a file-like object with the contents of the file
        # This can then be used to open the workbook from memory without writing anything to disk e.g.
        # - book = openpyxl.load_workbook(self.get_file())
        # - book = xlrd.open_workbook(file_contents=self.get_file().read())
        return io.BytesIO(self.data)


def transfer_comments(target,comment_source):
    # Format this AtomicaSpreadsheet based on the extra meta-content in comment_source
    #
    # In reality, a new spreadsheet is created with values from this AtomicaSpreadsheet
    # and cell-wise formatting from the comment_source AtomicaSpreadsheet. If a cell exists in
    # this spreadsheet and not in the source, it will be retained as-is. If more cells exist in
    # the comment_source than in this spreadsheet, those cells will be dropped. If a sheet exists in
    # the comment_source and not in the current workbook, it will be added

    assert isinstance(comment_source, AtomicaSpreadsheet)

    this_workbook = openpyxl.load_workbook(target.get_file(),data_only=False) # This is the value source workbook
    old_workbook = openpyxl.load_workbook(comment_source.get_file(),data_only=False) # A openpyxl workbook for the old content

    for sheet in this_workbook.worksheets:

        # If this sheet isn't in the old workbook, do nothing
        if sheet.title not in old_workbook.sheetnames:
            continue

        # Transfer comments
        for row in old_workbook[sheet.title].rows:
            for cell in row:
                if cell.comment:
                    sheet[cell.coordinate].comment = Comment(cell.comment.text, '')

    f = io.BytesIO()
    this_workbook.save(f)
    f.flush()
    f.seek(0)
    target.data = f.read()

def read_tables(worksheet):
    # This function takes in a openpyxl worksheet, and returns tables
    # A table consists of a block of rows with any #ignore rows skipped
    # This function will start at the top of the worksheet, read rows into a buffer
    # until it gets to the first entirely empty row
    # And then returns the contents of that buffer as a table. So a table is a list of openpyxl rows
    # This function continues until it has exhausted all of the rows in the sheet

    buffer = []
    tables = []
    for row in worksheet.rows:

        # Skip any rows starting with '#ignore'
        if row[0].value and row[0].value.startswith('#ignore'):
            continue  # Move on to the next row if row skipping is marked True

        # Find out whether we need to add the row to the buffer
        for cell in row:
            if cell.value:  # If the row has a non-empty cell, add the row to the buffer
                buffer.append(row)
                break
        else: # If the row was empty, then yield the buffer and flag that it should be cleared at the next iteration
            if buffer:
                tables.append(buffer) # Only append the buffer if it is not empty
            buffer = []

    # After the last row, if the buffer has some un-flushed contents, then yield it
    if buffer:
        tables.append(buffer)

    return tables


def write_matrix(worksheet,start_row,nodes,entries,formats,references=None, enable_diagonal=True, boolean_choice=False,widths=None):
    # This function writes a matrix
    # It gets used for
    # - Transfer matrix
    # - Interactions matrix
    # - Framework transition/link matrix
    #
    # - nodes is a list of strings used to label the rows and columns
    # - entries is a dict where where key is a tuple specifying (from,to) = (row,col) and
    # the value is the string to write to the matrix
    # - If 'enable_diagonal' is False, then the diagonal will be forced to be 'N.A.'. If an entry
    #   is specified for an entry on the diagonal and enable_diagonal=False, an error will be thrown
    # - boolean_choice is like namer/marker mode. If True, entries can only be Y/N based on the truthiness of the value in the entries dict
    #
    # table_references is a dict that contains a mapping between the tuple (to,from) and a cell. This can be
    # subsequently used to programatically block out time-dependent rows

    if not references:
        references = {x:x for x in nodes} # This is a null-mapping that takes say 'adults'->'adults' thus simplifying the workflow. Otherwise, it's assumed a reference exists for every node

    table_references = {}
    values_written = {}

    # Write the headers
    for i,node in enumerate(nodes):
        worksheet.write_formula(start_row+i+1, 0  , references[node], formats['center_bold'],value=node)
        update_widths(widths,0,node)
        worksheet.write_formula(start_row  , i+1, references[node], formats['center_bold'],value=node)
        update_widths(widths,i+1,node)

    # Prepare the content - first replace the dict with one keyed by index. This is because we cannot apply formatting
    # after writing content, so have to do the writing in a single pass over the entire matrix
    if boolean_choice:
        content = np.full((len(nodes),len(nodes)),'N',dtype=object) # This will also coerce the value to string in preparation for writing
    else:
        content = np.full((len(nodes),len(nodes)),'',dtype=object) # This will also coerce the value to string in preparation for writing

    for interaction, value in entries.items():
        from_node,to_node = interaction
        if not enable_diagonal and from_node == to_node:
            raise AtomicaException('Trying to write a diagonal entry to a table that is not allowed to contain diagonal terms') # This is because data loss will occur if the user adds entries on the diagonal, then writes the table, and then reads it back in
        from_idx = nodes.index(from_node)
        to_idx = nodes.index(to_node)
        if boolean_choice:
            value = 'Y' if value else 'N'
        content[from_idx,to_idx] = value

    # Write the content
    for from_idx in range(0,len(nodes)):
        for to_idx in range(0,len(nodes)):
            row = start_row+1+from_idx
            col = to_idx+1
            if not enable_diagonal and to_idx == from_idx: # Disable the diagonal if that's what's desired
                val = FS.DEFAULT_SYMBOL_INAPPLICABLE
                worksheet.write(row,col,val, formats["center"])
                worksheet.data_validation(xlrc(row,col), {"validate": "list", "source": ["N.A."]})
            else:
                val = content[from_idx,to_idx]
                worksheet.write(row,col, content[from_idx,to_idx], formats["center_unlocked"])
                if boolean_choice:
                    worksheet.data_validation(xlrc(row,col), {"validate": "list", "source": ["Y","N"]})
                    worksheet.conditional_format(xlrc(row,col), {'type': 'cell','criteria':'equal to','value':'"Y"','format':formats['unlocked_boolean_true']})
                    worksheet.conditional_format(xlrc(row,col), {'type': 'cell','criteria':'equal to','value':'"N"','format':formats['unlocked_boolean_false']})
            table_references[(nodes[from_idx],nodes[to_idx])] = xlrc(row,col,True,True) # Store reference to this interaction
            values_written[table_references[(nodes[from_idx],nodes[to_idx])]] = val

    next_row = start_row + 1 + len(nodes) + 1
    return next_row,table_references,values_written


class TimeDependentConnections(object):
    # A TimeDependentConnection structure is suitable when there are time dependent interactions between two quantities
    # This class is used for transfers and interactions
    # The content that it writes consists of
    # - A connection matrix table that has Y/N selection of which interactions are present between two things
    # - A set of pairwise connections specifying to, from, units, assumption, and time
    # Interactions can have a diagonal, whereas transfers cannot (e.g. a population can infect itself but cannot transfer to itself)

    def __init__(self, code_name, full_name, tvec, pops, type, ts=None):
        # INPUTS
        # - code_name -
        # - full_name - the name of this quantity e.g. 'Aging'
        # - tvec - time values for the time-dependent rows
        # - pops - list of strings to use as the rows and columns - these are typically lists of population code names
        # - ts - all of the non-empty TimeSeries objects used. An interaction can only be Y/N for clarity, if it is Y then
        #   a row is displayed for the TimeSeries. Actually, the Y/N can be decided in the first instance based on the provided TimeSeries i.e.
        #   if a TimeSeries is provided for an interaction, then the interaction must have been marked with Y
        # type - 'transfer' or 'interaction'. A transfer cannot have diagonal entries, and can have Number or Probability formats. An Interaction can have
        # diagonal entries and only has N.A. formats
        self.code_name = code_name
        self.full_name = full_name
        self.type = type
        self.pops = pops
        self.tvec = tvec
        self.ts = ts if ts is not None else sc.odict()

        if self.type == 'transfer':
            self.enable_diagonal = False
            self.allowed_units = [FS.QUANTITY_TYPE_NUMBER.title(), FS.QUANTITY_TYPE_PROBABILITY.title()]
        elif self.type == 'interaction':
            self.enable_diagonal = True
            self.allowed_units = [FS.DEFAULT_SYMBOL_INAPPLICABLE.title()]
        else:
            raise AtomicaException('Unknown TimeDependentConnections type - must be "transfer" or "interaction"')

    def __repr__(self):
        return '<TDC %s "%s">' % (self.type.title(),self.code_name)

    @staticmethod
    def from_tables(tables,interaction_type):
        # interaction_type is 'transfer' or 'interaction'
        from .structure import TimeSeries # Import here to avoid circular reference

        # Read the names
        code_name = tables[0][1][0].value
        full_name = tables[0][1][1].value
        interaction_type = interaction_type

        # Read the pops
        pops = [x.value for x in tables[1][0][1:] if x.value]
        # TODO - At the moment, the Y/N content of the table depends on what timeseries is provided
        # Therefore, we only need to read in the TimeSeries to tell whether or not a connection exists
        # The convention is that the connection will be read if the TO and FROM pop match something in the pop list
        tvec = np.array([x.value for x in tables[2][0][6:]],dtype=float) # The 6 matches the offset in write() below
        ts_entries = sc.odict()
        for row in tables[2][1:]:
            if row[0].value != '...':
                assert row[0].value in pops, 'Population "%s" not found - should be contained in %s' % (row[0].value, pops)
                assert row[2].value in pops, 'Population "%s" not found - should be contained in %s' % (row[2].value, pops)
                vals = [x.value for x in row]
                from_pop = vals[0]
                to_pop = vals[2]
                units = vals[3].lower().strip() if vals[3] else None
                assumption = vals[4] # This is the assumption cell
                assert vals[5] == 'OR' # Double check we are reading a time-dependent row with the expected shape
                ts = TimeSeries(format=units,units=units)
                if assumption:
                    ts.insert(None, assumption)
                for t, v in zip(tvec, vals[6:]):
                    if v is not None:
                        ts.insert(t, v)
                ts_entries[(from_pop,to_pop)] = ts

        if not ts_entries:
            logger.warning('TDC "%s" did not contain any data - is this intentional?' % (full_name))

        return TimeDependentConnections(code_name, full_name, tvec, pops, interaction_type, ts=ts_entries)

    def write(self,worksheet,start_row,formats,references=None, widths = None):

        if not references:
            references = {x:x for x in self.pops} # Default null mapping for populations

        ### First, write the titles
        current_row = start_row
        worksheet.write(current_row, 0, 'Abbreviation', formats["center_bold"])
        update_widths(widths, 0, 'Abbreviation')
        worksheet.write(current_row, 1, 'Full Name', formats["center_bold"])
        update_widths(widths, 1, 'Full Name')

        current_row += 1
        worksheet.write(current_row, 0, self.code_name)
        update_widths(widths, 0, self.code_name)
        worksheet.write(current_row, 1, self.full_name)
        references[self.code_name] = "='%s'!%s" % (worksheet.name, xlrc(current_row, 0, True, True))
        references[self.full_name] = "='%s'!%s" % (worksheet.name, xlrc(current_row, 1, True, True))  # Reference to the full name

        ### Then, write the matrix
        current_row += 2 # Leave a blank row below the matrix

        # Note - table_references are local to this TimeDependentConnections instance
        # For example, there could be two transfers, and each of them could potentially transfer between 0-4 and 5-14
        # so the worksheet might contain two references from 0-4 to 5-14 but they would be for different transfers and thus
        # the time-dependent rows would depend on different boolean table cells
        current_row,table_references,values_written = write_matrix(worksheet,current_row,self.pops,self.ts,formats,references,enable_diagonal=self.enable_diagonal,boolean_choice=True,widths=widths)

        ### Finally, write the time dependent part
        headings = []
        headings.append('') # From
        headings.append('') # --->
        headings.append('') # To
        headings.append('Units')
        headings.append('Constant')
        headings.append('') # OR
        headings += [float(x) for x in self.tvec] # Times
        for i, entry in enumerate(headings):
            worksheet.write(current_row, i, entry, formats['center_bold'])
            update_widths(widths,i,entry)

        # Now, we will write a wrapper that gates the content
        # If the gating cell is 'Y', then the content will be displayed, otherwise not
        def gate_content(content,gating_cell):
            if content.startswith('='): # If this is itself a reference
                return ('=IF(%s="Y",%s,"...")' % (gating_cell, content[1:]))
            else:
                return('=IF(%s="Y","%s","...")' % (gating_cell,content))

        for from_idx in range(0,len(self.pops)):
            for to_idx in range(0, len(self.pops)):
                current_row += 1
                from_pop = self.pops[from_idx]
                to_pop = self.pops[to_idx]
                entry_tuple = (from_pop,to_pop)
                entry_cell = table_references[entry_tuple]

                if entry_tuple in self.ts:
                    ts = self.ts[entry_tuple]
                    format = formats['not_required']
                else:
                    ts = None
                    format = formats['unlocked']

                if ts:
                    worksheet.write_formula(current_row, 0, gate_content(references[from_pop], entry_cell), formats['center_bold'], value=from_pop)
                    update_widths(widths, 0, from_pop)
                    worksheet.write_formula(current_row, 1, gate_content('--->', entry_cell), formats['center'], value='--->')
                    worksheet.write_formula(current_row, 2, gate_content(references[to_pop], entry_cell), formats['center_bold'], value=to_pop)
                    update_widths(widths, 2, to_pop)
                    worksheet.write(current_row, 3, ts.format.title())
                    update_widths(widths, 3, ts.format.title())

                    if self.allowed_units:
                        worksheet.data_validation(xlrc(current_row, 3), {"validate": "list", "source": self.allowed_units})
                    worksheet.write(current_row, 4, ts.assumption, format)
                    worksheet.write_formula(current_row, 5, gate_content('OR', entry_cell), formats['center'], value='OR')
                    # update_widths(widths, 5,  '...') # The largest length it will be here is '...' so use that

                else:
                    worksheet.write_formula(current_row, 0, gate_content(references[from_pop], entry_cell), formats['center_bold'], value='...')
                    worksheet.write_formula(current_row, 1, gate_content('--->', entry_cell), formats['center'], value='...')
                    worksheet.write_formula(current_row, 2, gate_content(references[to_pop], entry_cell), formats['center_bold'], value='...')
                    worksheet.write_blank(current_row, 3, '')
                    if self.allowed_units:
                        worksheet.data_validation(xlrc(current_row, 3), {"validate": "list", "source": self.allowed_units})
                    worksheet.write_blank(current_row, 4, '', format)
                    worksheet.write_formula(current_row, 5, gate_content('OR', entry_cell), formats['center'], value='...')
                    # update_widths(widths, 5,  '...')

                # Write hyperlink - it's a bit convoluted because we can't read back the contents of the original cell to know
                # whether it was originally Y or N
                if values_written[entry_cell] != FS.DEFAULT_SYMBOL_INAPPLICABLE:
                    worksheet.write_url(entry_cell, 'internal:%s!%s' % (worksheet.name, xlrc(current_row, 2)),cell_format=formats['center_unlocked'],string=values_written[entry_cell])

                offset = 6  # The time values start in this column (zero based index)
                content = [None]*len(self.tvec)

                if ts:
                    for t, v in zip(ts.t, ts.vals):
                        idx = np.where(self.tvec == t)[0][0]
                        content[idx] = v

                for idx, v in enumerate(content):
                    if v is None:
                        worksheet.write_blank(current_row, offset + idx, v, format)
                    else:
                        worksheet.write(current_row, offset + idx, v, format)

                # Conditional formatting for the assumption, depending on whether time-values were entered
                fcn_empty_times = 'COUNTIF(%s:%s,"<>" & "")>0' % (xlrc(current_row, offset), xlrc(current_row, offset + idx))
                worksheet.conditional_format(xlrc(current_row, 4), {'type': 'formula', 'criteria': '=' + fcn_empty_times, 'format': formats['ignored']})
                worksheet.conditional_format(xlrc(current_row, 4), {'type': 'formula', 'criteria': '=AND(%s,NOT(ISBLANK(%s)))' % (fcn_empty_times, xlrc(current_row, 4)), 'format': formats['ignored_warning']})

                # Conditional formatting for the row - it has a white background if the gating cell is 'N'
                worksheet.conditional_format('%s' % (xlrc(current_row, 4)), {'type': 'formula', 'criteria': '=%s<>"Y"' % (entry_cell), 'format': formats['white_bg']})
                worksheet.conditional_format('%s:%s' % (xlrc(current_row, offset), xlrc(current_row, offset + idx)), {'type': 'formula', 'criteria': '=%s<>"Y"' % (entry_cell), 'format': formats['white_bg']})

        current_row += 2

        return current_row


class TimeDependentValuesEntry(object):
    """
    Template table requesting time-dependent values, with each instantiation iterating over an item type.
    Argument 'value_attribute' specifies which attribute within item specs should contain the parsed values.
    If argument 'iterate_over_links' is True, table rows are actually for links between items of the iterated type.
    Self connections are not included by default, but can be turned on by an optional argument.
    """

    # A TDVE table is used for representing Characteristics and Parameters that appear in the Parset, a quantity
    # that has one sparse time array for each population. A TDVE table contains
    # - An ordered list of TimeSeries objects
    # - A name for the quantity (as this is what gets printed and read, it's usually a full name rather than a code name)
    # - Optionally a list of allowed units - All TimeSeries objects must have units contained in this list
    # - A time axis (e.g. np.arange(2000,2019)) - all TimeSeries time values must exactly match one of the values here
    #   i.e. you cannot try to write a TimeSeries that has a time value that doesn't appear as a table heading

    def __init__(self, name, tvec, ts = None, allowed_units = None):
        # ts - An odict where the key is a population name and the value is a TimeSeries
        # name - This is the name of the quantity i.e. the full name of the characteristic or parameter
        # tvec - The time values that will be written in the headings
        # allowed_units - Possible values for the unit selection dropdown
        if ts is None:
            ts = sc.odict()

        self.name = name
        self.tvec = tvec
        self.ts = ts
        self.allowed_units = allowed_units

    def __repr__(self):
        output= sc.desc(self)
        return output

    @property
    def has_data(self):
        # Returns True if any of the time series
        return any([x.has_data for x in self.ts.values()])

    @staticmethod
    def from_rows(rows):
        # Given a set of openpyxl rows, instantiate a TimeDependentValuesEntry object
        # That is, the parent object e.g. Databook() is responsible for finding where the TDVE table is,
        # and reading all of the rows associated with it (skipping #ignored rows) and then passing those rows,
        # unparsed, to this function
        #
        # 'units', 'uncertainty', and 'constant' are optional - if 'constant' is present then expect
        # that the column after it contains 'or'

        from .structure import TimeSeries # Import here to avoid circular reference

        # First, read the headings
        vals = [x.value for x in rows[0]]
        name = vals[0].strip()

        lowered_headings = [x.lower() if isinstance(x,string_types) else x for x in vals]

        # We can optionally have units, uncertainty, and constant
        # nb. finding the index means this is robust to extra empty
        # columns, a user adding one of the these fields to a single table on a page
        # might introduce a blank column to all of the other TDVE elements on the page too
        # so the code below should be able to deal with this
        if 'units' in lowered_headings:
            units_index = lowered_headings.index('units')
        else:
            units_index = None

        if 'uncertainty' in lowered_headings:
            uncertainty_index = lowered_headings.index('uncertainty')
        else:
            uncertainty_index = None

        if 'constant' in lowered_headings:
            constant_index = lowered_headings.index('constant')
        elif 'assumption' in lowered_headings:
            constant_index = lowered_headings.index('assumption')
        else:
            constant_index = None

        offset = 1 + (1 if units_index else 0) + (1 if uncertainty_index else 0) + (2 if constant_index else 0) # Index where the time values start
        tvec = np.array(vals[offset:],dtype=float)
        ts_entries = sc.odict()

        # For each TimeSeries that we will instantiate
        for row in rows[1:]:
            vals = [x.value for x in row]
            series_name = vals[0]

            if units_index is not None:
                units = vals[units_index].lower().strip() if vals[units_index] else None
                format = units
            else:
                units = None
                format = None

            ts = TimeSeries(format=format,units=units)

            if uncertainty_index is not None:
                sigma = vals[uncertainty_index]
                if sigma is not None and sigma != FS.DEFAULT_SYMBOL_INAPPLICABLE.title():
                    ts.sigma = float(sigma)
            else:
                ts.sigma = None

            if constant_index is not None:
                constant = vals[constant_index]
                if constant is not None and constant != FS.DEFAULT_SYMBOL_INAPPLICABLE.title():
                    ts.assumption = float(constant)
            else:
                ts.assumption = None

            if constant_index is not None:
                assert vals[offset - 1] == 'OR'  # Check row is as expected

            data = vals[offset:]

            for t,v in zip(tvec,data):
                if np.isfinite(t) and v is not None: # Ignore any times that are NaN - this happens if the cell was empty and casted to a float
                    ts.insert(t,v)
            ts_entries[series_name] = ts

        tvec = tvec[np.isfinite(tvec)] # Remove empty entries from the array
        return TimeDependentValuesEntry(name,tvec,ts_entries)

    def write(self,worksheet,start_row,formats,references=None,widths=None,assumption_heading='Constant',write_units=True,write_uncertainty=False,write_assumption=True):
        # references is a dict where the key is a string value and the content is a cell
        # Any populations that appear in this dict will have their value replaced by a reference
        # formats should be the dict returned by `excel.standard_formats` when it was called to add
        # formatting to the Workbook containing the worksheet passed in here.
        #
        # widths should be a dict that will store sizing information for some of the columns
        # it is updated in place
        # - assumption_heading : This is the string heading for the 'Constant'/'Assumption' (constant in databook, assumption in progbook)

        if not references:
            references = dict()

        current_row = start_row

        # First, assemble and write the headings
        headings = []
        headings.append(self.name)
        offset = 1 # This is the column where the time values start

        if write_units:
            headings.append('Units')
            units_index = offset # Column to write the units in
            offset += 1

        if write_uncertainty:
            headings.append('Uncertainty')
            uncertainty_index = offset  # Column to write the units in
            offset += 1

        if write_assumption:
            headings.append(assumption_heading)
            headings.append('')
            constant_index = offset
            offset += 2

        headings += [float(x) for x in self.tvec]
        for i,entry in enumerate(headings):
            worksheet.write(current_row, i, entry, formats['bold'])
            update_widths(widths,i,entry)

        # Now, write the TimeSeries objects - self.ts is an odict and whatever pops are present will be written in whatever order they are in
        for row_name, row_ts in self.ts.items():
            current_row += 1

            # Write the name
            if row_name in references:
                worksheet.write_formula(current_row, 0, references[row_name], formats['center_bold'],value=row_name)
                update_widths(widths, 0, row_name)
            else:
                worksheet.write_string(current_row, 0, row_name, formats['center_bold'])
                update_widths(widths, 0, row_name)

            # Write the units
            if write_units:
                worksheet.write(current_row,units_index,row_ts.format.title() if row_ts.format else None)
                update_widths(widths, units_index, row_ts.format.title() if row_ts.format else None)

                if self.allowed_units: # Add validation if a list of options is specified
                    worksheet.data_validation(xlrc(current_row, units_index),{"validate": "list", "source": self.allowed_units})

            if write_uncertainty:
                if row_ts.sigma is None:
                    worksheet.write(current_row,uncertainty_index, row_ts.sigma,formats['unlocked'])
                else:
                    worksheet.write(current_row,uncertainty_index,row_ts.sigma,formats['not_required'])

            if row_ts.has_data:
                format = formats['not_required']
            else:
                format = formats['unlocked']

            if write_assumption:
                # Write the assumption
                worksheet.write(current_row,constant_index,row_ts.assumption, format)
                # Write the separator between the assumptions and the time values
                worksheet.write(current_row,constant_index+1,'OR',formats['center'])
                update_widths(widths, constant_index+1, 'OR')

            # Write the time values
            content = [None]*len(self.tvec)

            for t,v in zip(row_ts.t,row_ts.vals):
                idx = np.where(self.tvec == t)[0][0] # If this fails there must be a (forbidden) mismatch between the TimeSeries and the Databook tvec
                content[idx] = v

            for idx,v in enumerate(content):
                if v is None:
                    worksheet.write_blank(current_row, offset+idx, v, format)
                else:
                    worksheet.write(current_row, offset+idx, v, format)

            # Conditional formatting for the assumption
            fcn_empty_times = 'COUNTIF(%s:%s,"<>" & "")>0' % (xlrc(current_row,offset),xlrc(current_row,offset+idx))
            # Hatched out if the cell will be ignored
            worksheet.conditional_format(xlrc(current_row, 2), {'type': 'formula', 'criteria':'='+fcn_empty_times,'format':formats['ignored']})
            worksheet.conditional_format(xlrc(current_row, 2), {'type': 'formula', 'criteria':'=AND(%s,NOT(ISBLANK(%s)))' % (fcn_empty_times,xlrc(current_row,2)),'format':formats['ignored_warning']})


        return current_row+2 # Add two so there is a blank line after this table



class ProgramEntry(object):
    ''' Class for generating the program data entry spreadsheet '''

    def __init__(self, sheet=None, first_row=None, first_col=None, content=None, row_interval=None):
        ''' Initialize '''

        self.sheet = sheet
        self.first_row = first_row
        self.first_col =  first_col if first_col is not None else 2
        self.content = content
        self.row_interval = row_interval if row_interval is not None else 3

        # Set the data range
        self.first_data_row = first_row + 2
        self.num_data_rows = len(self.content["row_names"])
        if self.content["row_levels"] is not None:
            self.first_col += 1
            self.num_data_rows *= len(self.content["row_levels"])
            self.num_data_rows += len(self.content["row_names"]) - 1

        self.first_data_col = first_col if first_col is not None else 2
        self.num_data_cols = len(self.content["column_names"])

        self.last_data_row = self.first_data_row + self.num_data_rows - 1
        self.last_data_col = self.first_data_col + self.num_data_cols - 1

        self.start = xlrc(self.first_data_row, self.first_data_col)
        self.end = xlrc(self.last_data_row, self.last_data_col) 


    def __repr__(self):
        ''' Representation '''
        output= sc.desc(self)
        return output


    def param_refs(self):
        ''' Returns a list of references to the program names '''
        prog_range = range(self.first_data_row, self.last_data_row+1)
        
        range_vals = []
        for row in prog_range: 
            cell_address = xlrc(row, self.first_data_col, row_abs=True, col_abs=True)
            range_vals.append("='%s'!%s" % (self.sheet.get_name(), cell_address))
        return range_vals


    def emit(self, formats, data=None, rc_row_align='right', rc_title_align='right', widths=None):  # only important for row/col titles
        """ Emits the range and returns the new current row in the given sheet """
        
        # Top-of-page headers
        self.formats = formats
        self.sheet.write(self.first_row, 0, self.content["name"], formats['bold'])
        
        # Write the main heading for the assumption column
        if self.content["assumption"] and self.first_row == 0 and self.content["assumption_properties"]['title'] is not None:
            self.sheet.write(self.first_row, self.last_data_col+2, self.content["assumption_properties"]['title'], self.formats['rc_title']['left']['F'])

        # Write the heading for the other columns
        for i, name in enumerate(self.content["column_names"]):
            self.sheet.write(self.first_row+1, self.first_data_col+i, name, formats['rc_title']['right']['T'])

        # Write the subheading for the assumption column(s)
        if self.content["assumption"]:
            for index, col_name in enumerate(self.content["assumption_properties"]['columns']):
                self.sheet.write(self.first_row+1, self.last_data_col+2+index, col_name, formats['rc_title']['right']['T'])

        # Initialise first row and row levels
        current_row = self.first_data_row
        num_levels = len(self.content["row_levels"]) if self.content["row_levels"] is not None else 1
        
        # Get row names and formats
        if self.content["row_levels"] is None:
            row_names = [[name] for name in self.content["row_names"]]
            row_formats = [self.content["row_format"] for name in self.content["row_names"]]
        else:
            row_names = [[name, level] for name in self.content["row_names"] for level in self.content["row_levels"]]
            if self.content["row_formats"] is not None:
                row_formats = [row_format for name in self.content["row_names"] for row_format in self.content["row_formats"]]
            else:
                row_formats = [self.content["row_format"] for name in self.content["row_names"] for level in self.content["row_levels"]]

        # Iterate over rows, incrementing current_row as we go
        for i, names_format in enumerate(zip(row_names, row_formats)):
            names, row_format = names_format
            start_col = self.first_data_col - len(names)
            
            # Write row name(s)
            for n, name in enumerate(names):
                self.sheet.write(current_row, start_col+n, name, formats['rc_title']['left']['T'])
                update_widths(widths,start_col+n,name)

            # Write first part (data) if present
            savedata = False
            if self.content["data"]:
                try:
                    for j, item in enumerate(self.content["data"][i]):
                        self.sheet.write(current_row, self.first_data_col+j, item, self.formats[row_format])
                    savedata = True  # It saved successfully
                except:
                    errormsg = 'WARNING, failed to save "%s" with data:\n%s' % (self.content["name"], self.content["data"])
                    raise AtomicaException(errormsg)
                    savedata = False
            if not savedata:
                for j in range(self.num_data_cols):
                    self.sheet.write(current_row, self.first_data_col+j, None, self.formats[row_format])

            # Write second part (assumption or program effects) if present
            if self.content["assumption"]:

                self.sheet.write(current_row, self.last_data_col + 1,
                                     self.content["assumption_properties"]['connector'],
                                     formats['center_bold'])

                for index, col_name in enumerate(self.content["assumption_properties"]['columns']):
                    saveassumptiondata = False
                    if self.content["assumption_data"]:
                        try:
                            self.sheet.write(current_row, self.last_data_col+2+index, self.content["assumption_data"][i][index], self.formats[row_format])
                            saveassumptiondata = True
                        except Exception as E:
                            errormsg = 'WARNING, failed to save assumption "%s" with data:\n%s\nError message:\n (%s)' % (
                                self.content["name"], self.content["assumption_data"], repr(E))
                            saveassumptiondata = False
                            raise AtomicaException(errormsg)
                    if not saveassumptiondata:
                        self.sheet.write(current_row, self.last_data_col+2+index, None, self.formats[row_format])

            # Insert any data validations
            if self.content["validation"]:
                for col,source in self.content["validation"].iteritems():
                    self.sheet.data_validation(xlrc(current_row,col), {"validate": "list", "source": source})

            current_row += 1
            if num_levels > 1 and ((i + 1) % num_levels) == 0:  # shift between the blocks
                current_row += 1
        # Return the new current_row plus spacing
        return current_row + self.row_interval # for spacing

